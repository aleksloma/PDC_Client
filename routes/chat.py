"""Client-side chat endpoints — SSE stream + sidebar helpers + B2C-only stubs.

The chat-stream endpoint matches the B2C `/api/chat/{chat_id}/chat/stream`
shape that `dashboard.js` consumes:
  - emits `data: {...}\\n\\n` JSON events
  - `{progress: true, message: "..."}` for status text in the loading bubble
  - `{partial: true, answer, image_base64, chart_n, chart_total, conv_id}`
    for per-chart partial results (not used here — enterprise build emits
    a single final event), kept in the contract for parity
  - `{done: true, answer, image_base64, table, conv_id, tokens}` final event
  - `{error: "..."}` on failure (kill-switch surfaces here)

LLM work is done in a worker thread; results come back through an
`asyncio.Queue` (same pattern as the B2C `chat_stream_api`).
"""
from __future__ import annotations

import asyncio
import io
import json
import re
import secrets
import threading
import time
from concurrent.futures import ThreadPoolExecutor

from fastapi import APIRouter, File, Request, UploadFile
from fastapi.responses import JSONResponse, Response, StreamingResponse

import local_store
import run_chat_local
import brain_client
from brain_client import TenantRevokedError, BrainError, BrainTimeoutError
from logger_utils import log_with_sid
from schema_builder import _detect_language as _detect_lang_for_title


# Token-keyed in-memory cache of full result tables (the chat stream returns
# only a preview when `full_table_key` is set; the frontend fetches the full
# table via /api/chat/{chat_id}/full_table/{key}). Bounded LRU so very long
# sessions don't blow memory.
_FULL_TABLE_CACHE: dict[str, dict] = {}
_FULL_TABLE_ORDER: list[str] = []
_FULL_TABLE_MAX = 256

def _cache_full_table(table: dict) -> str:
    key = secrets.token_hex(8)
    _FULL_TABLE_CACHE[key] = table
    _FULL_TABLE_ORDER.append(key)
    while len(_FULL_TABLE_ORDER) > _FULL_TABLE_MAX:
        old = _FULL_TABLE_ORDER.pop(0)
        _FULL_TABLE_CACHE.pop(old, None)
    return key


# Durable full-table persistence (port of the B2C `download_full_excel`
# mechanism). A tabular RESULT is written to disk as
# `chatdata/{chat_id}/conversations/full/{key}.json` = {columns, rows, code,
# total_rows}, so "Download Excel" / "Show full table" survive a container
# restart AND can re-execute the stored code to return the COMPLETE (uncapped)
# result — not the 50-row preview. The in-memory LRU is kept as a fast path and
# as the fallback for chart-data ("Show data") records, which have no
# re-executable DataFrame code. Article V: local disk under DATA_ROOT only.
_FULL_KEY_RE = re.compile(r"[0-9a-fA-F]{16}")


def _persist_full_table(store, table: dict, code: str | None,
                        result_key: str | None = None) -> str | None:
    """Persist {columns, rows, code, total_rows} to disk and mirror it into the
    in-memory LRU. Returns the durable key, or None on failure (caller then
    leaves `full_table_key` unset and the frontend exports the preview rows).

    `result_key`: set for one table of a multi-table answer — names the dict
    entry of the re-executed RESULT this table comes from (see
    `_reexecute_full_df`)."""
    try:
        key = secrets.token_hex(8)  # 16 hex chars — matches _FULL_KEY_RE
        record = _json_safe({
            "columns": table.get("columns", []),
            "rows": table.get("rows", []),
            "total_rows": table.get("total_rows"),
        })
        if code:
            record["code"] = code
        if result_key is not None:
            record["result_key"] = result_key
        full_dir = store.conversations_dir / "full"
        full_dir.mkdir(parents=True, exist_ok=True)
        (full_dir / f"{key}.json").write_text(
            json.dumps(record, ensure_ascii=False), encoding="utf-8")
        # Fast-path / fallback copy in the bounded LRU.
        _FULL_TABLE_CACHE[key] = record
        _FULL_TABLE_ORDER.append(key)
        while len(_FULL_TABLE_ORDER) > _FULL_TABLE_MAX:
            old = _FULL_TABLE_ORDER.pop(0)
            _FULL_TABLE_CACHE.pop(old, None)
        return key
    except Exception as e:
        log_with_sid(getattr(store, "chat_id", ""), "warning",
                     f"PERSIST_FULL_TABLE_FAILED: {e}")
        return None


def _load_full_table_record(store, key: str) -> dict | None:
    """Resolve a full-table / chart-data record by key: durable disk first
    (survives restart, may carry re-executable `code`), then the in-memory LRU
    (chart-data "Show data", no code). Returns None when the key is malformed or
    unknown. The regex guard also blocks path traversal (Article VII)."""
    if not (key and _FULL_KEY_RE.fullmatch(key)):
        return None
    try:
        p = store.conversations_dir / "full" / f"{key}.json"
        if p.exists():
            return json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        log_with_sid(getattr(store, "chat_id", ""), "warning",
                     f"FULL_TABLE_READ_FAILED: {e}")
    return _FULL_TABLE_CACHE.get(key)


async def _reexecute_full_df(chat_id: str, code: str | None, result_key: str | None = None,
                             *, drop_df_keys=None):
    """Re-run stored code locally to produce the COMPLETE (uncapped) DataFrame.
    Returns a DataFrame or None (caller falls back to the stored preview rows).

    `result_key`: for multi-table answers (RESULT was a dict of DataFrames) the
    durable record stores which dict entry this table came from — pick it out
    of the re-executed dict before the normal normalization.

    `drop_df_keys`: df keys removed from the exec namespace AFTER load (role
    gate defense in depth — the dashboard tile-refresh path passes the denied
    set; the full_table/Download-Excel callers pass nothing and stay ungated
    by design: viewing existing data is never blocked retroactively). Filtered
    post-load on purpose — _load_dataframes_cached's cache key is per-chat,
    not per-user, so the cache must always hold the full set.

    Article II: execution is local — nothing here touches the brain.
    Article V: reads local DataFrames only. Article IV: never raises."""
    if not code:
        return None
    try:
        import pandas as pd
        from code_exec import safe_execute
        store = local_store.ChatDataStore(chat_id)
        loop = asyncio.get_running_loop()
        dfs = await loop.run_in_executor(_EXEC, store.load_dataframes)
        if drop_df_keys:
            dfs = {k: v for k, v in dfs.items() if k not in drop_df_keys}
        if not dfs:
            return None
        exec_out = await loop.run_in_executor(
            _EXEC, lambda: safe_execute(code, dfs, sid=chat_id))
        if not isinstance(exec_out, dict) or exec_out.get("error"):
            return None
        obj = exec_out.get("result")
        if isinstance(obj, dict) and result_key is not None:
            obj = obj.get(result_key)
        # _normalize_df_for_table: same pivot normalization as the chat-time
        # serializer (index surfaced, MultiIndex columns flattened) so the
        # full-table view / Excel download match what the chat rendered (QA 2.2)
        from run_chat_local import _normalize_df_for_table
        if isinstance(obj, pd.DataFrame):
            return _normalize_df_for_table(obj)
        if isinstance(obj, pd.Series):
            return _normalize_df_for_table(obj.reset_index())
        if hasattr(obj, "data") and isinstance(obj.data, pd.DataFrame):
            return _normalize_df_for_table(obj.data)
        return None
    except Exception as e:
        log_with_sid(chat_id, "warning", f"DOWNLOAD_REEXEC_FAILED: {e}")
        return None


def _auto_analysis_busy_response(store, email: str, chat_id: str):
    """409 when Auto Analytics is running for this chat, else None (QA 2.3).

    While the job runs, its 4 workers occupy the single code-exec worker and
    the shared brain connection pool, so an interactive question times out
    ~90s later with a misleading "temporarily unavailable" error. Telling the
    user honestly and immediately — BEFORE any history append — is the fix.
    Fail-open: a corrupt meta must never block chat (Article IV)."""
    try:
        import auto_analytics as aa
        if aa.get_auto_analysis_state(store.read_meta()).get("status") == "processing":
            log_with_sid(email, "info", f"CHAT_BLOCKED_AUTO_ANALYSIS chat={chat_id}")
            return JSONResponse({
                "error": ("Auto Analytics is currently running for this chat. "
                          "Your question can start as soon as it finishes — "
                          "please ask again then."),
                "busy": "auto_analysis",
            }, status_code=409)
    except Exception as e:
        log_with_sid(email, "warning", f"AUTO_BUSY_CHECK_FAILED chat={chat_id}: {e}")
    return None


# In-progress generation registry. A conv_id present here has a worker thread
# generating right now. Used ONLY by the lightweight status endpoint so a page
# reloaded/reopened mid-generation can show the working indicator and block new
# questions until the AI turn is persisted. It NEVER affects persistence — the
# worker remains the sole, exactly-once persister.
_INPROGRESS_CONVS: set[str] = set()
_INPROGRESS_LOCK = threading.Lock()


def _mark_generating(conv_id: str) -> None:
    try:
        with _INPROGRESS_LOCK:
            _INPROGRESS_CONVS.add(conv_id)
    except Exception:
        pass


def _unmark_generating(conv_id: str) -> None:
    try:
        with _INPROGRESS_LOCK:
            _INPROGRESS_CONVS.discard(conv_id)
    except Exception:
        pass


def _is_generating(conv_id: str) -> bool:
    try:
        with _INPROGRESS_LOCK:
            return conv_id in _INPROGRESS_CONVS
    except Exception:
        return False


# Cooperative-cancel registry (STOP button). A conv_id here has had a stop
# requested; the worker checks it between charts and halts at the next chart
# boundary (the in-flight chart still finishes — cancellation is cooperative,
# not instant). Shares the in-progress lock. Never affects exactly-once
# persistence — the worker still persists whatever charts it produced.
_CANCEL_CONVS: set[str] = set()


def _request_cancel(conv_id: str) -> None:
    try:
        with _INPROGRESS_LOCK:
            _CANCEL_CONVS.add(conv_id)
    except Exception:
        pass


def _is_cancelled(conv_id: str) -> bool:
    try:
        with _INPROGRESS_LOCK:
            return conv_id in _CANCEL_CONVS
    except Exception:
        return False


def _clear_cancel(conv_id: str) -> None:
    try:
        with _INPROGRESS_LOCK:
            _CANCEL_CONVS.discard(conv_id)
    except Exception:
        pass


router = APIRouter(prefix="/api/chat", tags=["client-chat"])

# Worker pool for the SSE per-request thread (Article VI)
_EXEC = ThreadPoolExecutor(max_workers=4, thread_name_prefix="client_chat")

import atexit
atexit.register(lambda: _EXEC.shutdown(wait=False, cancel_futures=True))


# One implementation for SSE payloads AND persistence — the full pandas/numpy
# normalizer (Timestamp→ISO, NaT→None, catch-all str()) lives in local_store
# next to the history writers that depend on it.
_json_safe = local_store._json_safe


# Max serialized size of a chart's source data we will PERSIST into history so
# the "Show data" button survives reload. Oversize or missing data is omitted
# (the button just won't appear for that chart) — we never bloat the on-disk
# conversation or risk a serialization failure.
_PERSIST_CHART_DATA_MAX_CHARS = 200_000


def _persistable_chart_data(cd):
    """Return JSON-safe chart_data if it is usable and within the size cap,
    else None. Mirrors the live-cache guard (a dict with `rows` or `tables`)."""
    try:
        if not (isinstance(cd, dict) and (cd.get("rows") or cd.get("tables"))):
            return None
        safe = _json_safe(cd)
        if len(json.dumps(safe, ensure_ascii=False)) > _PERSIST_CHART_DATA_MAX_CHARS:
            return None
        return safe
    except Exception:
        return None


def _build_chart_entry(c: dict) -> dict:
    """One persisted chart entry for a multi-chart turn: image + answer plus its
    OWN code and (size-capped) chart_data, so reload restores Show code / Show
    data per chart (PART B)."""
    entry = {"image_base64": c.get("image_base64"), "answer": c.get("answer", "")}
    if c.get("code"):
        entry["code"] = c["code"]
    cd = _persistable_chart_data(c.get("chart_data"))
    if cd is not None:
        entry["chart_data"] = cd
    return entry


def _build_ai_history_record(err_msg, single, combined_answer, combined_codes,
                             usage, charts, full_table_key=None,
                             full_table_keys=None, combined_tables=None):
    """Build the AI-turn history record persisted EXACTLY ONCE by the worker.

    Returns None when there is nothing to persist (never writes an empty turn).
    Enriches with PER-CHART code + chart_data (PART B). Backward-compatible
    shape: 0 imgs → image_base64=None; 1 img → image_base64 + record-level
    code/chart_data; 2+ imgs → images=[{image_base64, answer, code?, chart_data?}].
    `full_table_key` (single-shot tabular result) is persisted so a reloaded
    conversation can re-execute for the FULL Download Excel / Show full table.
    """
    if err_msg:
        return {"role": "ai", "content": err_msg, "ts": time.time()}
    if single is not None:
        rec = {
            "role": "ai",
            "content": single.get("text", ""),
            "image_base64": single.get("image_base64"),
            "table": single.get("table"),
            "code": single.get("code"),
            "usage": single.get("usage"),
            "ts": time.time(),
        }
        if full_table_key:
            rec["full_table_key"] = full_table_key
        # Multi-table answer: persist the tables array (mirrors the multi-chart
        # `images` array) + per-table durable keys aligned by index.
        tables = single.get("tables")
        if tables:
            rec["tables"] = tables
            if full_table_keys:
                rec["full_table_keys"] = full_table_keys
        cd = _persistable_chart_data(single.get("chart_data"))
        if cd is not None:
            rec["chart_data"] = cd
        return rec
    if combined_answer is not None:
        imgs = [c for c in charts if c.get("image_base64")]
        rec = {
            "role": "ai",
            "content": combined_answer,
            "image_base64": None,
            "table": None,
            "code": "\n\n###NEXT_PLOT###\n\n".join(combined_codes or []),
            "usage": usage or {},
            "ts": time.time(),
        }
        if len(imgs) >= 2:
            rec["images"] = [_build_chart_entry(c) for c in imgs]
        elif len(imgs) == 1:
            c = imgs[0]
            rec["image_base64"] = c.get("image_base64")
            if c.get("code"):
                rec["code"] = c["code"]
            cd = _persistable_chart_data(c.get("chart_data"))
            if cd is not None:
                rec["chart_data"] = cd
        # Mixed dashboard answer: the KPI/table blocks' tables ride alongside
        # the charts (per-table durable keys aligned by index).
        if combined_tables:
            rec["tables"] = combined_tables
            if full_table_keys:
                rec["full_table_keys"] = full_table_keys
        return rec
    return None


def _build_stopped_record(combined_answer, combined_codes, usage, charts):
    """Persist a STOPPED AI turn (user clicked Stop) — never the planner
    NO_CODE fallback or a late single-shot result.

    ≥1 chart produced → the partial charts (same per-chart code/chart_data
    shape as a normal multi-chart turn) with a stopped marker appended to the
    content, plus a `stopped: True` flag. 0 charts → a short "Response stopped
    by user." turn. Returns a record dict (never None — a stop always yields a
    turn so reload shows it consistently). Reuses _build_chart_entry /
    _persistable_chart_data so per-chart Show data/code survive reload.
    """
    note = "Response stopped by user."
    imgs = [c for c in charts if c.get("image_base64")]
    if imgs:
        answers = [c.get("answer", "") for c in imgs if c.get("answer")]
        base = "\n\n".join(a for a in answers if a)
        content = (base + "\n\n⏹ " + note).strip() if base else ("⏹ " + note)
        rec = {
            "role": "ai",
            "content": content,
            "image_base64": None,
            "table": None,
            "code": "\n\n###NEXT_PLOT###\n\n".join(combined_codes or []),
            "usage": usage or {},
            "stopped": True,
            "ts": time.time(),
        }
        if len(imgs) >= 2:
            rec["images"] = [_build_chart_entry(c) for c in imgs]
        elif len(imgs) == 1:
            c = imgs[0]
            rec["image_base64"] = c.get("image_base64")
            if c.get("code"):
                rec["code"] = c["code"]
            cd = _persistable_chart_data(c.get("chart_data"))
            if cd is not None:
                rec["chart_data"] = cd
        return rec
    return {
        "role": "ai",
        "content": note,
        "stopped": True,
        "ts": time.time(),
    }


def _require_chat(request: Request, chat_id: str):
    email = request.session.get("email")
    if not email:
        return None, JSONResponse({"error": "Not authenticated"}, status_code=401)
    if not local_store.chat_exists(chat_id):
        return None, JSONResponse({"error": "Chat not found"}, status_code=404)
    owner = local_store.get_chat_meta_owner(chat_id)
    if owner == email:
        return email, None
    # Allow shared recipients to access the chat (chat-level + conversation-level
    # sharing both populate meta.json["sharing"]["shared_with"]).
    try:
        store = local_store.ChatDataStore(chat_id)
        sharing = (store.read_meta().get("sharing") or {}).get("shared_with") or []
        if email in [s.lower() for s in sharing]:
            return email, None
    except Exception:
        pass
    return None, JSONResponse({"error": "Access denied"}, status_code=403)


# ---------------------------------------------------------------------------
# Welcome + schema + history
# ---------------------------------------------------------------------------
@router.get("/{chat_id}/welcome")
async def welcome(request: Request, chat_id: str):
    """Same response shape as the B2C `/api/chat/{chat_id}/welcome`:
    {message, language, suggested_questions} — `dashboard.js` reads
    `welcomeData.message` and `welcomeData.suggested_questions`.
    """
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    store = local_store.ChatDataStore(chat_id)
    meta = store.read_meta()
    msg = meta.get("welcome_message") or store.get_welcome() or ""
    questions = meta.get("suggested_questions") or store.get_suggested_questions() or []
    # Detect language from file descriptions (same logic as global welcome handler)
    from schema_builder import _detect_language
    descs = [f.get("file_description", "") for f in meta.get("files", []) if f.get("file_description")]
    lang = _detect_language(" ".join(descs)) if descs else "en"
    return {
        "message": msg,
        "language": lang,
        "suggested_questions": questions,
    }


@router.get("/{chat_id}/schema")
async def get_schema(request: Request, chat_id: str):
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    store = local_store.ChatDataStore(chat_id)
    meta = store.read_meta()
    # "Data as of" for chats that use registered database tables. Additive:
    # file-only chats get db_tables=[] and data_as_of=None (UI renders
    # nothing). refreshed_at prefers the REGISTRY (authoritative) over the
    # chat meta's display cache; a table gone from the registry or whose
    # snapshot file vanished reports missing=true (load_dataframes skips that
    # key, and the frontend key-freeze already disables items referencing it).
    db_tables = []
    data_as_of = None
    db_entries = local_store.db_entries_from_meta(meta)
    if db_entries:
        try:
            from db_sources import DataSourceStore
            registry = {t.get("id"): t for t in DataSourceStore().list_tables()}
        except Exception:
            registry = {}
        # Role gate (additive `allowed` flag so the UI can grey refresh
        # buttons proactively). Helper failure → None → every row reports
        # allowed=true (the server-side refresh gate is the enforcement).
        allowed_ids = None
        try:
            import roles_store
            allowed_ids = roles_store.allowed_table_ids_for(email)
        except Exception as e:
            log_with_sid(email, "warning", f"SCHEMA_ROLE_PROBE_FAILED: {e}",
                         chat_id=chat_id)
        for entry in db_entries:
            db = entry.get("db") or {}
            tid = db.get("table_id")
            reg = registry.get(tid)
            refreshed = (reg or {}).get("refreshed_at") or db.get("refreshed_at")
            missing = True
            try:
                missing = reg is None or not local_store.db_snapshot_path(tid).exists()
            except Exception as e:
                log_with_sid(email, "warning",
                             f"DB_TABLE_MISSING_PROBE_FAILED table={tid}: {e}",
                             chat_id=chat_id)
            db_tables.append({
                "df_key": entry.get("file_name"),
                "table_id": tid,
                "display_name": db.get("display_name") or entry.get("file_name"),
                "is_connector": bool(db.get("is_connector")),
                "auto_included": bool(db.get("auto_included")),
                "row_count": (reg or {}).get("row_count") or db.get("row_count"),
                "refreshed_at": refreshed,
                "missing": missing,
                "allowed": bool(db.get("is_connector")) or allowed_ids is None
                           or tid in allowed_ids,
            })
        stamps = [t["refreshed_at"] for t in db_tables if t.get("refreshed_at")]
        data_as_of = min(stamps) if stamps else None  # oldest data in the chat
    return {
        "chat_id": chat_id,
        "files": meta.get("files", []),
        "common_fields": meta.get("common_fields", []),
        "db_tables": db_tables,
        "data_as_of": data_as_of,
    }


@router.get("/{chat_id}/file_fingerprints")
async def file_fingerprints(request: Request, chat_id: str):
    """Identity of this chat's SOURCE files for the Add Data name-collision
    dialog: {files: {source_filename: {size_bytes, sha256}}}. The browser
    compares sizes first and hashes the selected File locally only on a size
    tie — everything stays client↔client, nothing reaches the brain
    (Article II). Hashing is disk work → executor (never on the event loop).
    Article IV: any failure degrades to an empty/partial map — the frontend
    then falls back to name-only collision handling."""
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    import hashlib

    store = local_store.ChatDataStore(chat_id)

    def _compute() -> dict:
        out: dict = {}
        for fp in sorted(store.files_dir.iterdir()):
            if not (fp.is_file() and not fp.name.startswith(".")):
                continue
            # The name is always reported — a hash/stat failure degrades to a
            # name-only entry (nulls), which the frontend treats as "contents
            # unknown → show the dialog", never as "no collision".
            entry: dict = {"size_bytes": None, "sha256": None}
            try:
                entry["size_bytes"] = fp.stat().st_size
                h = hashlib.sha256()
                with fp.open("rb") as fh:
                    for chunk in iter(lambda: fh.read(1024 * 1024), b""):
                        h.update(chunk)
                entry["sha256"] = h.hexdigest()
            except Exception as e:
                log_with_sid(chat_id, "warning",
                             f"FINGERPRINT_FAILED {fp.name}: {e}")
            out[fp.name] = entry
        return out

    try:
        loop = asyncio.get_running_loop()
        files = await loop.run_in_executor(_EXEC, _compute)
    except Exception as e:
        log_with_sid(chat_id, "error", f"FINGERPRINTS_ERROR: {e}")
        files = {}
    return {"files": files}


def _probe_structure(data: bytes, filename: str) -> dict | None:
    """Fast, best-effort header-level structure of a tabular file WITHOUT the
    full detection pipeline: {sheet_name_or_'': [column strings]}.

    - .xlsx/.xlsm — openpyxl read_only: VISIBLE sheets, first non-empty row
      within the first 50 rows per sheet as approximate columns.
    - .csv/.tsv — first non-empty row within the first 50 lines (exact header).
    Returns None for unsupported formats or any parse failure — callers must
    then skip the comparison (Article IV: never raise)."""
    import io as _io
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    try:
        if ext in ("xlsx", "xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
            out: dict = {}
            try:
                for ws in wb.worksheets:
                    if getattr(ws, "sheet_state", "visible") != "visible":
                        continue
                    header: list = []
                    for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
                        cells = ["" if v is None else str(v).strip() for v in row]
                        while cells and not cells[-1]:
                            cells.pop()
                        if any(cells):
                            header = cells
                            break
                    out[ws.title] = header
            finally:
                wb.close()
            return out
        if ext in ("csv", "tsv"):
            import csv as _csv
            text = data.decode("utf-8-sig", errors="replace")
            reader = _csv.reader(_io.StringIO(text),
                                 delimiter="\t" if ext == "tsv" else ",")
            for i, row in enumerate(reader):
                if i >= 50:
                    break
                cells = [str(v).strip() for v in row]
                while cells and not cells[-1]:
                    cells.pop()
                if any(cells):
                    return {"": cells}
            return {"": []}
        return None
    except Exception:
        return None


@router.post("/{chat_id}/probe_columns")
async def probe_columns(request: Request, chat_id: str,
                        file: UploadFile = File(...)):
    """Header-level structure comparison for the Add Data collision dialog:
    the uploaded file vs the chat's EXISTING same-named file, WITHOUT running
    the detection pipeline (capped at the first 50 rows per sheet — fast).
    Client-container only; nothing reaches the brain (Article II). Cell values
    are returned to the browser but never logged (filename + match only).
    Any failure → {ok: false} — the dialog keeps its generic warning."""
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    try:
        from pathlib import Path as _Path
        fname = _Path(file.filename or "").name   # strip any path components
        if not fname:
            return {"ok": False}
        store = local_store.ChatDataStore(chat_id)
        existing_path = store.files_dir / fname
        if not existing_path.exists():
            return {"ok": False}
        data = await file.read()
        loop = asyncio.get_running_loop()
        uploaded = await loop.run_in_executor(_EXEC, _probe_structure, data, fname)
        existing_bytes = await loop.run_in_executor(_EXEC, existing_path.read_bytes)
        existing = await loop.run_in_executor(_EXEC, _probe_structure, existing_bytes, fname)
        if uploaded is None or existing is None:
            return {"ok": False}
        match = uploaded == existing
        log_with_sid(chat_id, "info", f"PROBE_COLUMNS file={fname} match={match}")
        return {"ok": True, "match": match, "uploaded": uploaded, "existing": existing}
    except Exception as e:
        log_with_sid(chat_id, "warning", f"PROBE_COLUMNS_FAILED: {e}")
        return {"ok": False}


@router.post("/{chat_id}/schema")
async def save_schema(request: Request, chat_id: str):
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    body = await request.json()
    store = local_store.ChatDataStore(chat_id)
    meta = store.read_meta()
    # Body shape from dashboard.js: {files: [{file_name, fields:{...}, file_description}]}
    posted = {f.get("file_name"): f for f in (body.get("files") or []) if f.get("file_name")}
    for entry in meta.get("files", []):
        name = entry.get("file_name")
        if name in posted:
            p = posted[name]
            if "fields" in p:
                entry.setdefault("schema", {})["fields"] = p["fields"]
            if "file_description" in p:
                entry["file_description"] = p["file_description"]
    store.write_meta(meta)
    return {"ok": True}


@router.get("/{chat_id}/conversation/{conv_id}/history")
async def history(request: Request, chat_id: str, conv_id: str):
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    store = local_store.ChatDataStore(chat_id)
    return {"history": store.get_history(conv_id)}


@router.get("/{chat_id}/conversation/{conv_id}/status")
async def conversation_status(request: Request, chat_id: str, conv_id: str):
    """Is a generation worker currently running for this conversation?

    Registry lookup only (no I/O). Lets a page reloaded/reopened mid-generation
    show the working indicator and block new questions until the worker has
    persisted the AI turn. Mirrors the Auto Analytics status-poll pattern.
    """
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    return {"generating": _is_generating(conv_id)}


@router.post("/{chat_id}/conversation/{conv_id}/stop")
async def conversation_stop(request: Request, chat_id: str, conv_id: str):
    """Request cancellation of an in-progress generation for this conversation.

    Cooperative: the worker halts at the NEXT chart boundary (the in-flight
    chart still finishes), persists the partial AI turn (the charts produced so
    far, same shape as a normal turn), and emits a normal `done`. Idempotent —
    setting the flag when nothing is running is a harmless no-op.
    """
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    _request_cancel(conv_id)
    log_with_sid("stop", "info", "CHAT_STOP_REQUESTED", user=email,
                 chat_id=chat_id, conv_id=conv_id)
    return {"ok": True, "stopping": True}


@router.get("/{chat_id}/history")
async def chat_history(request: Request, chat_id: str):
    """Legacy single-conversation history — return the most recent conv."""
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    store = local_store.ChatDataStore(chat_id)
    # Pick the newest conversation file
    try:
        convs = sorted(store.conversations_dir.glob("*.jsonl"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not convs:
            return {"history": []}
        conv_id = convs[0].stem
        return {"history": store.get_history(conv_id), "conv_id": conv_id}
    except Exception:
        return {"history": []}


@router.post("/{chat_id}/deactivate")
async def deactivate(request: Request, chat_id: str):
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    local_store.AuthStore().deactivate_chat(email, chat_id)
    return {"ok": True}


# ---------------------------------------------------------------------------
# SSE chat stream  — the working chat path
# ---------------------------------------------------------------------------
@router.post("/{chat_id}/chat/stream")
async def chat_stream(request: Request, chat_id: str):
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    body = await request.json()
    question = (body.get("question") or "").strip()
    conv_id = body.get("conv_id")
    if not question:
        return JSONResponse({"error": "Question cannot be empty."}, status_code=400)

    store = local_store.ChatDataStore(chat_id)
    busy = _auto_analysis_busy_response(store, email, chat_id)
    if busy is not None:
        return busy
    # Off the event loop — parsing the source files blocks every other
    # request otherwise (same pattern as _reexecute_full_df / refresh_item).
    loop = asyncio.get_running_loop()
    dfs = await loop.run_in_executor(_EXEC, store.load_dataframes)
    if not dfs:
        return JSONResponse({"error": "Chat dataset is empty."}, status_code=400)
    schema_docs = await loop.run_in_executor(_EXEC, store.schema_docs)

    sid = secrets.token_hex(8)
    if not conv_id:
        conv_id = store.new_conversation(title=question[:80])
        local_store.AuthStore().record_conversation(email, chat_id, conv_id, title=question[:80])
        # Seed the conversation with the welcome message (matches B2C)
        welcome = store.get_welcome()
        if welcome:
            store.append_history(conv_id, {"role": "ai", "content": welcome, "ts": time.time()})

    history_rows = store.get_history(conv_id)
    store.append_history(conv_id, {"role": "human", "content": question, "ts": time.time()})

    log_with_sid(sid, "info", "CHAT_STREAM_REQ", user=email, chat_id=chat_id, conv_id=conv_id, q=question[:120])

    async def _sse_generator():
        loop = asyncio.get_running_loop()
        queue: asyncio.Queue = asyncio.Queue()

        def _worker():
            # PART A: accumulate the AI turn and persist it EXACTLY ONCE here, in
            # the worker thread, so a browser refresh/close (which cancels the SSE
            # consumer below) can never discard the generated response. The
            # consumer performs live UX only — it does NO persistence. The worker
            # runs to completion regardless of the connection.
            w_charts: list[dict] = []      # per-chart {image_base64, answer, code, chart_data}
            w_combined_answer = None        # set by the multi-plot done event
            w_combined_codes: list = []
            w_usage: dict = {}
            w_single = None                 # single-shot result dict
            w_full_table_key = None         # durable key for a tabular result
            w_full_table_keys = None        # per-table keys for a multi-table result
            w_combined_tables = None        # tables of a mixed charts+tables answer
            err_msg = None
            cancelled = False               # STOP requested mid-generation
            # Clear any stale cancel flag from a prior attempt, then mark this
            # conv in-progress for the duration of the worker so a page reloaded
            # mid-generation sees generating=true. Both flags are cleared in
            # `finally` AFTER persistence — once unmarked, the AI turn is readable.
            _clear_cancel(conv_id)
            _mark_generating(conv_id)
            try:
                # Use the multi-plot generator: it yields per-chart partials when
                # the planner emitted ###NEXT_PLOT### blocks, and a single
                # {single_response, result} for normal one-shot answers.
                gen = run_chat_local.run_chat_multi_plot(
                    sid=sid, dfs=dfs, schema_docs=schema_docs,
                    question=question, history_rows=history_rows,
                    user_email=email,
                )
                for event in gen:
                    # Durable full-table persistence for a single-shot tabular
                    # result: write {columns, rows, code, total_rows} to disk and
                    # stash the key ON the shared event dict BEFORE the consumer
                    # sees it — so the live download and the persisted history
                    # record share ONE durable key that re-executes for the FULL
                    # result (and survives a container restart / reload).
                    if isinstance(event, dict) and event.get("single_response"):
                        _res = event.get("result") or {}
                        _tbl = _res.get("table")
                        if isinstance(_tbl, dict) and _tbl.get("rows"):
                            _k = _persist_full_table(store, _tbl, _res.get("code"))
                            if _k:
                                event["full_table_key"] = _k
                                w_full_table_key = _k
                        # Multi-table answer (RESULT was a dict of DataFrames):
                        # one durable key per table, each tagged with the dict
                        # entry (`result_key`) it re-executes to.
                        _tbls = _res.get("tables")
                        if isinstance(_tbls, list) and _tbls:
                            _keys = []
                            for _t in _tbls:
                                _k = _persist_full_table(
                                    store, _t, _res.get("code"),
                                    result_key=_t.get("title"))
                                _keys.append(_k)
                            event["full_table_keys"] = _keys
                            w_full_table_keys = _keys
                    # Mixed dashboard done event (charts + KPI/table blocks):
                    # persist each table with its OWN block code + dict entry.
                    if (isinstance(event, dict) and event.get("done")
                            and not event.get("single_response") and event.get("tables")):
                        _tbls = event.get("tables") or []
                        _codes = event.get("table_codes") or []
                        _rkeys = event.get("table_result_keys") or []
                        _keys = []
                        for _i, _t in enumerate(_tbls):
                            _keys.append(_persist_full_table(
                                store, _t,
                                _codes[_i] if _i < len(_codes) else None,
                                result_key=_rkeys[_i] if _i < len(_rkeys) else None))
                        event["full_table_keys"] = _keys
                    # Stream live first (UX identical to before), then accumulate.
                    loop.call_soon_threadsafe(queue.put_nowait, ("event", event))
                    try:
                        if isinstance(event, dict):
                            if event.get("partial"):
                                w_usage = event.get("usage") or w_usage
                                if event.get("image_base64"):
                                    w_charts.append({
                                        "image_base64": event.get("image_base64"),
                                        "answer": event.get("answer", ""),
                                        "code": event.get("code"),
                                        "chart_data": event.get("chart_data"),
                                    })
                            elif event.get("done") and not event.get("single_response"):
                                w_combined_answer = event.get("combined_answer", "")
                                w_combined_codes = event.get("combined_codes") or []
                                w_usage = event.get("total_usage") or {}
                                if event.get("tables"):
                                    w_combined_tables = event.get("tables")
                                    w_full_table_keys = event.get("full_table_keys") or w_full_table_keys
                            elif event.get("single_response"):
                                w_single = event.get("result") or {}
                    except Exception:
                        pass
                    # Cooperative STOP: halt at this chart boundary (the chart just
                    # streamed is kept). No further charts are pulled from the gen.
                    if _is_cancelled(conv_id):
                        cancelled = True
                        log_with_sid(sid, "info", "CHAT_STOP_AT_BOUNDARY",
                                     charts=len(w_charts))
                        break
                # If STOP broke us out before the generator's own done event,
                # synthesize the combined result from the charts produced so far
                # (so persistence + the client's done handler behave normally) and
                # emit a done so the live stream finalizes and shows the partial.
                if cancelled and w_single is None and w_combined_answer is None:
                    if w_charts:
                        answers = [c.get("answer", "") for c in w_charts if c.get("answer")]
                        w_combined_answer = "\n\n".join(answers) if answers else "Analysis stopped."
                        w_combined_codes = [c.get("code") for c in w_charts if c.get("code")]
                    loop.call_soon_threadsafe(queue.put_nowait, ("event", {
                        "partial": False, "done": True,
                        "combined_answer": w_combined_answer or "",
                        "combined_codes": w_combined_codes,
                        "total_usage": w_usage,
                    }))
            except TenantRevokedError:
                err_msg = "Service unavailable. Please contact your administrator."
                loop.call_soon_threadsafe(queue.put_nowait, ("error", err_msg))
            except BrainTimeoutError:
                # BrainTimeoutError subclasses BrainError — caught FIRST. A
                # timeout under load (e.g. Auto Analytics occupying the shared
                # connection pool) is contention, not an outage; the old
                # "temporarily unavailable" wording was misleading (QA 2.3).
                err_msg = "The analysis service is busy right now. Please try again in a moment."
                loop.call_soon_threadsafe(queue.put_nowait, ("error", err_msg))
            except BrainError:
                err_msg = "Analysis service is temporarily unavailable."
                loop.call_soon_threadsafe(queue.put_nowait, ("error", err_msg))
            except Exception as e:
                log_with_sid(sid, "error", f"CHAT_THREAD_ERROR: {type(e).__name__}: {e}")
                err_msg = "Something went wrong while processing your request."
                loop.call_soon_threadsafe(queue.put_nowait, ("error", err_msg))

            # ── Exactly-once AI-turn persistence (Articles IV & V) ──
            try:
                if cancelled:
                    # User clicked Stop. Persist a STOPPED turn — the partial
                    # charts (if any) or a short "stopped by user" note — never
                    # the planner NO_CODE error or a late single-shot result.
                    record = _build_stopped_record(
                        w_combined_answer, w_combined_codes, w_usage, w_charts)
                else:
                    record = _build_ai_history_record(
                        err_msg, w_single, w_combined_answer, w_combined_codes,
                        w_usage, w_charts, full_table_key=w_full_table_key,
                        full_table_keys=w_full_table_keys,
                        combined_tables=w_combined_tables)
                if record is not None:
                    store.append_history(conv_id, record)
                    if not err_msg and not cancelled:
                        answer_for_title = (
                            (w_single or {}).get("text", "") if w_single is not None
                            else (w_combined_answer or ""))
                        try:
                            human_count = sum(
                                1 for m in store.get_history(conv_id)
                                if m.get("role") == "human")
                            if human_count == 2:
                                _start_title_generation(
                                    email, chat_id, conv_id, question, answer_for_title)
                        except Exception:
                            pass
            except Exception as e:
                log_with_sid(sid, "error", f"CHAT_PERSIST_ERROR: {type(e).__name__}: {e}")
            finally:
                # Unmark only after persistence above has completed, so a status
                # poll that now sees generating=false will find the AI turn saved.
                # Clear the cancel flag too (paired with the start-of-worker clear).
                _unmark_generating(conv_id)
                _clear_cancel(conv_id)
                loop.call_soon_threadsafe(queue.put_nowait, None)

        threading.Thread(target=_worker, daemon=True).start()

        # First event — empty progress + conv_id so the frontend can capture the conv_id
        yield f'data: {json.dumps({"progress": True, "message": "Working...", "conv_id": conv_id})}\n\n'

        # The consumer streams events to the browser and caches live chart_data;
        # it persists NOTHING (the worker owns persistence — exactly-once).
        while True:
            event = await queue.get()
            if event is None:
                break
            tag, payload = event
            if tag == "error":
                err_event = {"error": payload, "conv_id": conv_id, "done": True}
                yield f"data: {json.dumps(err_event, ensure_ascii=False)}\n\n"
                continue
            if tag == "event":
                ev = payload
                # ── Multi-plot: per-chart partial event (live UX only) ──
                if ev.get("partial"):
                    img = ev.get("image_base64")
                    chart_answer = ev.get("answer", "")
                    # Cache the chart's source data for the client-side
                    # "Show data" button — reuses the full_table cache + endpoint.
                    # Client-only display; nothing to brain. `chart_data` is a
                    # single {rows} table or a multi-subplot {tables:[...]}.
                    chart_data_key = None
                    cd = ev.get("chart_data")
                    if isinstance(cd, dict) and (cd.get("rows") or cd.get("tables")):
                        chart_data_key = _cache_full_table(cd)
                    partial_payload = {
                        "partial": True,
                        "answer": chart_answer,
                        "image_base64": img,
                        "chart_n": ev.get("chart_n"),
                        "chart_total": ev.get("chart_total"),
                        "conv_id": conv_id,
                        "code": ev.get("code"),
                        "chart_data_key": chart_data_key,
                        "tokens": ev.get("usage") or {},
                    }
                    yield f"data: {json.dumps(_json_safe(partial_payload), ensure_ascii=False)}\n\n"
                    if img:
                        try:
                            brain_client.post_activity("plot_generated", email,
                                                       {"chat_id": chat_id, "conv_id": conv_id,
                                                        "chart_n": ev.get("chart_n")})
                        except Exception:
                            pass
                    continue

                # ── Multi-plot: final done event (live UX only) ──
                if ev.get("done") and not ev.get("single_response"):
                    answer = ev.get("combined_answer", "")
                    codes = ev.get("combined_codes") or []
                    usage = ev.get("total_usage") or {}
                    out = {
                        "done": True, "partial": False,
                        "conv_id": conv_id,
                        "answer": answer,
                        "image_base64": None, "table": None,
                        # Mixed dashboard answers carry the KPI/table blocks'
                        # results; the worker already stamped per-table keys.
                        "tables": ev.get("tables"),
                        "full_table_keys": ev.get("full_table_keys"),
                        "code": "\n\n###NEXT_PLOT###\n\n".join(codes),
                        "tokens": usage,
                    }
                    yield f"data: {json.dumps(_json_safe(out), ensure_ascii=False)}\n\n"
                    continue

                # ── Single-shot path: {single_response: True, result: ...} ─
                result = ev.get("result") or {}
                if result.get("image_base64"):
                    try:
                        brain_client.post_activity("plot_generated", email,
                                                   {"chat_id": chat_id, "conv_id": conv_id})
                    except Exception:
                        pass
                # The worker already persisted the tabular result to disk and
                # stashed its durable key on the event (re-executes for the FULL
                # result + survives restart). Use it; do not re-cache the preview.
                full_table_key = ev.get("full_table_key")
                # Chart source data for the "Show data" button — reuses the
                # full_table cache + endpoint. Client-only display. Single
                # {rows} table or multi-subplot {tables:[...]}.
                chart_data_key = None
                cd = result.get("chart_data")
                if isinstance(cd, dict) and (cd.get("rows") or cd.get("tables")):
                    chart_data_key = _cache_full_table(cd)
                out = {
                    "done": True,
                    "partial": False,
                    "conv_id": conv_id,
                    "answer": result.get("text", ""),
                    "image_base64": result.get("image_base64"),
                    "table": result.get("table"),
                    "tables": result.get("tables"),
                    "full_table_key": full_table_key,
                    "full_table_keys": ev.get("full_table_keys"),
                    "chart_data_key": chart_data_key,
                    "code": result.get("code"),
                    "tokens": result.get("usage") or {},
                }
                yield f"data: {json.dumps(_json_safe(out), ensure_ascii=False)}\n\n"

    return StreamingResponse(_sse_generator(), media_type="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
    })


# ---------------------------------------------------------------------------
# Edit-regenerate — verbatim port of global backend/routes/chat.py
# `edit_regenerate_api`, adapted for the brain/client split (LLM via brain,
# raw data + execution on client). dashboard.js triggers this from the
# pencil-edit affordance on a past user message.
# ---------------------------------------------------------------------------
@router.post("/{chat_id}/edit-regenerate")
async def edit_regenerate(request: Request, chat_id: str):
    """Edit the last user message and regenerate the AI response.

    Body: {edited_question: str, conv_id: str}
    Returns: {answer, image_base64?, table?, full_table_key?, conv_id, tokens}
    (or, for multi-chart responses, {answer, images: [...], conv_id, tokens}).

    Mirrors global's truncate-then-rerun behavior:
      1. Truncate the conv history to drop the last human turn (and everything
         after it).
      2. Append the edited question as the new human turn.
      3. Run `run_chat_multi_plot` against the local dfs (matches the normal
         chat-stream path), persisting the AI turn in the same JSONL shape.
    """
    email, err = _require_chat(request, chat_id)
    if err:
        return err

    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    edited_question = (body.get("edited_question") or "").strip()
    conv_id = (body.get("conv_id") or "").strip()
    if not edited_question:
        return JSONResponse({"error": "Question cannot be empty."}, status_code=400)
    if not conv_id:
        return JSONResponse({"error": "Conversation ID is required."}, status_code=400)

    store = local_store.ChatDataStore(chat_id)
    if not store.root.exists():
        return JSONResponse({"error": "Chat not found."}, status_code=404)
    busy = _auto_analysis_busy_response(store, email, chat_id)
    if busy is not None:
        return busy

    history = store.get_history(conv_id)
    if len(history) < 2:
        return JSONResponse(
            {"error": "Not enough messages to edit. Need at least one exchange."},
            status_code=400,
        )

    # Find the last human turn → keep everything *before* it.
    last_human_idx = None
    for i in range(len(history) - 1, -1, -1):
        if history[i].get("role") == "human":
            last_human_idx = i
            break
    if last_human_idx is None:
        return JSONResponse({"error": "No user message found to edit."}, status_code=400)

    store.truncate_conv_history(conv_id, last_human_idx)
    log_with_sid(chat_id, "info",
                 f"EDIT_REGENERATE user={email} conv_id={conv_id} truncated_to={last_human_idx}")

    try:
        loop = asyncio.get_running_loop()
        dfs = await loop.run_in_executor(_EXEC, store.load_dataframes)
        if not dfs:
            return JSONResponse({"error": "Chat dataset is empty."}, status_code=400)
        schema_docs = await loop.run_in_executor(_EXEC, store.schema_docs)

        sid = secrets.token_hex(8)
        history_rows = store.get_history(conv_id)
        # Append the edited human turn after truncation, before running.
        store.append_history(conv_id, {
            "role": "human", "content": edited_question, "ts": time.time(),
        })

        def _run_blocking():
            return list(run_chat_local.run_chat_multi_plot(
                sid=sid, dfs=dfs, schema_docs=schema_docs,
                question=edited_question, history_rows=history_rows,
                user_email=email,
            ))

        events = await loop.run_in_executor(_EXEC, _run_blocking)
    except TenantRevokedError:
        msg = "Service unavailable. Please contact your administrator."
        store.append_history(conv_id, {"role": "ai", "content": msg, "ts": time.time()})
        return JSONResponse({"error": msg, "conv_id": conv_id}, status_code=503)
    except BrainTimeoutError as e:
        # subclass of BrainError — caught first; see the chat_stream worker (QA 2.3)
        msg = "The analysis service is busy right now. Please try again in a moment."
        log_with_sid(chat_id, "warning", f"EDIT_REGENERATE_BRAIN_TIMEOUT: {e}")
        store.append_history(conv_id, {"role": "ai", "content": msg, "ts": time.time()})
        return JSONResponse({"error": msg, "conv_id": conv_id}, status_code=503)
    except BrainError as e:
        msg = "Analysis service is temporarily unavailable."
        log_with_sid(chat_id, "warning", f"EDIT_REGENERATE_BRAIN_ERROR: {e}")
        store.append_history(conv_id, {"role": "ai", "content": msg, "ts": time.time()})
        return JSONResponse({"error": msg, "conv_id": conv_id}, status_code=503)
    except Exception as e:
        log_with_sid(chat_id, "error", f"EDIT_REGENERATE_ERROR: {type(e).__name__}: {e}")
        return JSONResponse(
            {"error": "Something went wrong while processing your request."},
            status_code=500,
        )

    # Accumulate multi-plot images (same shape used in the normal SSE path so
    # persistence + frontend rendering match exactly).
    all_images: list[str] = []
    all_answers: list[str] = []
    single_result: dict | None = None
    combined_codes: list[str] = []
    final_usage: dict = {}

    for ev in events:
        if ev.get("partial"):
            img = ev.get("image_base64")
            if img:
                all_images.append(img)
                all_answers.append(ev.get("answer", ""))
            continue
        if ev.get("done") and not ev.get("single_response"):
            combined_codes = ev.get("combined_codes") or []
            final_usage = ev.get("total_usage") or {}
            combined_answer = ev.get("combined_answer", "")
            history_obj: dict = {
                "role": "ai", "content": combined_answer,
                "image_base64": None, "table": None,
                "code": "\n\n###NEXT_PLOT###\n\n".join(combined_codes),
                "usage": final_usage, "ts": time.time(),
            }
            if len(all_images) >= 2:
                history_obj["images"] = [
                    {"image_base64": img, "answer": ans}
                    for img, ans in zip(all_images, all_answers) if img
                ]
            elif len(all_images) == 1:
                history_obj["image_base64"] = all_images[0]
            # Mixed dashboard answer: persist the KPI/table blocks' tables with
            # per-table durable keys (same as chat_stream's worker).
            tbls = ev.get("tables") or []
            if tbls:
                codes_per = ev.get("table_codes") or []
                rkeys = ev.get("table_result_keys") or []
                keys = [
                    _persist_full_table(store, t,
                                        codes_per[i] if i < len(codes_per) else None,
                                        result_key=rkeys[i] if i < len(rkeys) else None)
                    for i, t in enumerate(tbls)
                ]
                history_obj["tables"] = tbls
                history_obj["full_table_keys"] = keys
            store.append_history(conv_id, history_obj)
            out: dict = {
                "ok": True, "done": True, "conv_id": conv_id,
                "answer": combined_answer,
                "image_base64": None, "table": None,
                "code": "\n\n###NEXT_PLOT###\n\n".join(combined_codes),
                "tokens": final_usage,
            }
            if "images" in history_obj:
                out["images"] = history_obj["images"]
            elif history_obj.get("image_base64"):
                out["image_base64"] = history_obj["image_base64"]
            if tbls:
                out["tables"] = tbls
                out["full_table_keys"] = history_obj.get("full_table_keys")
            return JSONResponse(_json_safe(out))

        # Single-shot path
        if ev.get("single_response"):
            single_result = ev.get("result") or {}
            # Persist the tabular result durably (with code) BEFORE the history
            # record, so the key can be embedded → a reloaded conversation
            # re-executes for the FULL Download Excel / Show full table.
            full_table_key = None
            tbl = single_result.get("table")
            if isinstance(tbl, dict) and tbl.get("rows"):
                full_table_key = _persist_full_table(
                    store, tbl, single_result.get("code"))
            # Multi-table answer: one durable key per table (see chat_stream).
            full_table_keys = None
            tbls = single_result.get("tables")
            if isinstance(tbls, list) and tbls:
                full_table_keys = [
                    _persist_full_table(store, t, single_result.get("code"),
                                        result_key=t.get("title"))
                    for t in tbls
                ]
            ai_record = {
                "role": "ai",
                "content": single_result.get("text", ""),
                "image_base64": single_result.get("image_base64"),
                "table": single_result.get("table"),
                "code": single_result.get("code"),
                "usage": single_result.get("usage"),
                "ts": time.time(),
            }
            if full_table_key:
                ai_record["full_table_key"] = full_table_key
            if tbls:
                ai_record["tables"] = tbls
                if full_table_keys:
                    ai_record["full_table_keys"] = full_table_keys
            store.append_history(conv_id, ai_record)
            chart_data_key = None
            cd = single_result.get("chart_data")
            if isinstance(cd, dict) and (cd.get("rows") or cd.get("tables")):
                chart_data_key = _cache_full_table(cd)
            out = {
                "ok": True, "done": True, "conv_id": conv_id,
                "answer": single_result.get("text", ""),
                "image_base64": single_result.get("image_base64"),
                "table": single_result.get("table"),
                "tables": single_result.get("tables"),
                "full_table_key": full_table_key,
                "full_table_keys": full_table_keys,
                "chart_data_key": chart_data_key,
                "code": single_result.get("code"),
                "tokens": single_result.get("usage") or {},
            }
            return JSONResponse(_json_safe(out))

    # Should not reach here, but degrade safely.
    return JSONResponse(
        {"error": "No response was produced.", "conv_id": conv_id},
        status_code=500,
    )


# ---------------------------------------------------------------------------
# Filename generator (used by the report download button)
# ---------------------------------------------------------------------------
@router.post("/{chat_id}/generate_filename")
async def generate_filename(request: Request, chat_id: str):
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    title = (body.get("title") or "analysis_report").strip()
    import re as _re
    safe = _re.sub(r"[^a-zA-Z0-9_-]", "_", title)[:40] or "analysis_report"
    return {"filename": safe}


# ---------------------------------------------------------------------------
# Title generation — background, fired after 2nd human message
# ---------------------------------------------------------------------------
def _start_title_generation(email: str, chat_id: str, conv_id: str,
                             question: str, answer: str) -> None:
    """Spin a daemon thread to call brain /v1/title and rename the conversation."""
    def _worker():
        try:
            lang_code = _detect_lang_for_title((question + " " + answer)[:300])
            lang_name = {"ka": "Georgian", "ru": "Russian"}.get(lang_code, "English")
            rsp = brain_client.title(
                sid=f"title:{conv_id}", question=question, answer=answer,
                lang=lang_name, user_email=email,
            )
            new_title = (rsp.get("title") or "").strip()
            if new_title:
                local_store.AuthStore().rename_conversation(email, conv_id, new_title)
                log_with_sid(email, "info", f"CONV_TITLE_UPDATED conv_id={conv_id} title={new_title!r}")
        except Exception as e:
            log_with_sid(email, "warning", f"TITLE_GEN_FAILED: {e}")
    threading.Thread(target=_worker, daemon=True).start()


# ---------------------------------------------------------------------------
# Chat sharing — implemented (brain SMTP relay sends invites)
# ---------------------------------------------------------------------------
@router.get("/{chat_id}/share")
async def share_get(request: Request, chat_id: str):
    """Return the current chat-level sharing record."""
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    store = local_store.ChatDataStore(chat_id)
    meta = store.read_meta()
    sharing = meta.get("sharing") or {}
    return {
        "shared_with": sharing.get("shared_with") or [],
        "owner": email,
    }


@router.post("/{chat_id}/share")
async def share_post(request: Request, chat_id: str):
    """Body: {emails: ["a@x.com", ...], message?: "..."}.

    Adds the recipients to this chat's sharing list and asks the brain to
    SMTP-relay an invite email to each. The brain's SMTP relay uses this
    tenant's smtp_* config (set in the per-tenant admin page).
    """
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    body = await request.json()
    raw_emails = body.get("emails") or []
    if isinstance(raw_emails, str):
        raw_emails = [e.strip() for e in raw_emails.replace(",", "\n").splitlines() if e.strip()]
    recipients = []
    import re as _re
    _email_re = _re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
    for e in raw_emails:
        e = (e or "").strip().lower()
        if _email_re.match(e) and e != email:
            recipients.append(e)
    if not recipients:
        return JSONResponse({"error": "Provide at least one valid recipient email."}, status_code=400)
    message_text = (body.get("message") or "").strip()

    store = local_store.ChatDataStore(chat_id)
    meta = store.read_meta()
    sharing = meta.get("sharing") or {"shared_with": []}
    existing = set(sharing.get("shared_with") or [])
    new_recipients = [r for r in recipients if r not in existing]
    existing.update(new_recipients)
    sharing["shared_with"] = sorted(existing)
    meta["sharing"] = sharing
    store.write_meta(meta)

    chat_title = meta.get("title", "")
    smtp_result = {"smtp_configured": False, "sent": [], "failed": []}
    if new_recipients:
        try:
            smtp_result = brain_client.send_share_email(
                to=new_recipients, subject=f"{email} shared an analysis with you",
                sender_email=email, chat_title=chat_title, message=message_text,
            ) or smtp_result
        except (TenantRevokedError, BrainError) as e:
            log_with_sid(email, "warning", f"SHARE_EMAIL_BRAIN_ERROR: {e}")
        except Exception as e:
            log_with_sid(email, "warning", f"SHARE_EMAIL_ERROR: {e}")

    return {
        "ok": True,
        "shared_with": sharing["shared_with"],
        "added": new_recipients,
        "email_sent": bool(smtp_result.get("sent")),
        "smtp_configured": bool(smtp_result.get("smtp_configured")),
        "failed": smtp_result.get("failed") or [],
    }


# ---------------------------------------------------------------------------
# B2C-only features — disabled in enterprise (clean errors so the UI handles gracefully)
# ---------------------------------------------------------------------------
def _disabled(message: str, code: int = 400):
    return JSONResponse({"error": message}, status_code=code)


@router.post("/{chat_id}/publish")
@router.post("/{chat_id}/unpublish")
async def publish_disabled(request: Request, chat_id: str):
    return _disabled("Public publish is not available in the on-prem build.")


@router.get("/{chat_id}/publish-status")
async def publish_status_disabled(request: Request, chat_id: str):
    return {"public": False}


@router.post("/{chat_id}/auto_analysis/start")
async def auto_analysis_start(request: Request, chat_id: str):
    """Kick off the auto-analysis background job. Returns immediately; poll
    `/auto_analysis/status` for state and finally fetch the PPTX from
    `/auto_analysis/download` when status is `done`."""
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    import auto_analytics as aa
    store = local_store.ChatDataStore(chat_id)
    state = aa.get_auto_analysis_state(store.read_meta())
    if state.get("status") == "processing":
        return {"ok": True, "status": "processing", "message": "Already running."}
    from datetime import datetime, timezone as _tz
    aa._update_state(
        store, status="processing",
        started_at=datetime.now(_tz.utc).isoformat(),
        finished_at=None, error=None, progress="Starting…", pptx_path=None,
    )
    aa.start_job(chat_id, email)
    log_with_sid(email, "info", f"AUTO_ANALYTICS_START chat={chat_id}")
    return {"ok": True, "status": "processing"}


@router.get("/{chat_id}/auto_analysis/status")
async def auto_analysis_status(request: Request, chat_id: str):
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    import auto_analytics as aa
    store = local_store.ChatDataStore(chat_id)
    return aa.get_auto_analysis_state(store.read_meta())


@router.get("/{chat_id}/auto_analysis/download")
async def auto_analysis_download(request: Request, chat_id: str):
    """Stream the auto-analysis PPTX. 404 until the job is `done`."""
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    import auto_analytics as aa
    store = local_store.ChatDataStore(chat_id)
    state = aa.get_auto_analysis_state(store.read_meta())
    if state.get("status") != "done":
        return JSONResponse(
            {"error": "Auto-analysis is not ready", "status": state.get("status")},
            status_code=404,
        )
    pptx_path = store.root / aa.AUTO_ANALYSIS_PPTX_NAME
    if not pptx_path.exists():
        return JSONResponse({"error": "Deck file is missing"}, status_code=404)
    from fastapi.responses import FileResponse
    return FileResponse(
        str(pptx_path),
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        filename=f"auto_analysis_{chat_id[:8]}.pptx",
    )


@router.get("/{chat_id}/suggested-questions")
async def suggested_questions_stub(request: Request, chat_id: str):
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    store = local_store.ChatDataStore(chat_id)
    return {"questions": store.get_suggested_questions()}


@router.get("/{chat_id}/full_table/{key}")
async def full_table_get(request: Request, chat_id: str, key: str):
    """Return the FULL row set for this key ("Show full table"). The chat stream
    sets `full_table_key` on responses that contain a tabular result; the
    frontend fetches the full rows via this endpoint.

    Re-executes the stored code locally to return the complete (uncapped)
    result, falling back to the stored preview rows if re-execution yields
    nothing (e.g. chart-data records, which carry no re-executable code)."""
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    store = local_store.ChatDataStore(chat_id)
    rec = _load_full_table_record(store, key)
    if not rec:
        return JSONResponse({"error": "Full table not found or expired."}, status_code=404)
    df = await _reexecute_full_df(chat_id, rec.get("code"), rec.get("result_key"))
    if df is not None and not df.empty:
        return _json_safe({
            "columns": list(df.columns),
            "rows": df.to_dict(orient="records"),
            "total_rows": int(len(df)),
        })
    return _json_safe(rec)


# --- Per-chart / per-table refresh -------------------------------------------
# Re-runs ONE item's stored code against the chat's CURRENT dataframes (after
# an Add Data update) and returns the re-rendered chart/table. Purely local
# code re-execution (Article II — nothing touches the brain), same execution
# path _reexecute_full_df uses. Body: {code, kind: "chart"|"table"}.
# Auth/permission/validation failures use HTTP status codes; EXECUTION failures
# return 200 {ok: false, error} so the frontend keeps the previous render and
# shows a small non-blocking note (Article IV: logged, safe fallback).

_DF_KEY_RE = re.compile(r"dfs\[\s*['\"]([^'\"]+)['\"]\s*\]")


def _role_refresh_block(email: str, chat_id: str, code: str):
    """Role gate for the refresh paths (chat refresh_item + dashboard tile
    refresh, both branches). Returns (denied_df_keys, blocked_display_names):

      denied_df_keys — frozenset of the chat's df keys whose backing DB table
        the REQUESTER's role does not cover (non-connector entries only —
        connectors are exempt by design). Passed as drop_df_keys so denied
        frames never enter the exec namespace even when unreferenced.
      blocked_display_names — the denied tables this item's code actually
        references (the same dfs['…'] key regex dashboard.js freezes on) —
        non-empty means the refresh must be refused; empty means the item
        only touches allowed tables and proceeds (per-table semantics).

    Genuine denials fail CLOSED by construction (.get() defaults resolve to
    Base → empty grant set). An unexpected crash inside the helper fails OPEN
    (logged ROLE_GATE_FAILED) — a bug here must not freeze every refresh
    fleet-wide for data the user can already see as a snapshot."""
    try:
        meta = local_store.ChatDataStore(chat_id).read_meta()
        entries = [e for e in local_store.db_entries_from_meta(meta)
                   if not (e.get("db") or {}).get("is_connector")]
        if not entries:
            return frozenset(), []
        import roles_store
        allowed = roles_store.allowed_table_ids_for(email)
        denied = {e["file_name"]: ((e.get("db") or {}).get("display_name") or e["file_name"])
                  for e in entries
                  if e.get("file_name")
                  and (e.get("db") or {}).get("table_id") not in allowed}
        if not denied:
            return frozenset(), []
        referenced = set(_DF_KEY_RE.findall(code or ""))
        return frozenset(denied), sorted(denied[k] for k in referenced if k in denied)
    except Exception as e:
        log_with_sid(email, "warning", f"ROLE_GATE_FAILED chat={chat_id}: {e}")
        return frozenset(), []


async def run_item_refresh(chat_id: str, code: str, kind: str, sid: str,
                           *, drop_df_keys=None) -> dict:
    """Re-run ONE item's stored code against the chat's current dataframes and
    return the re-rendered item. Shared by `refresh_item` (per-message refresh
    in /lab) and the dashboard tile refresh (routes/dashboards.py). Purely
    local (Article II). Returns `{ok: True, ...}` payloads identical to the
    historical refresh_item contract, or `{ok: False, error}` on any execution
    failure (Article IV: logged, safe fallback — never raises).

    `drop_df_keys` — see _reexecute_full_df: role-denied df keys filtered out
    AFTER the (per-chat, user-agnostic) cached load."""
    store = local_store.ChatDataStore(chat_id)
    loop = asyncio.get_running_loop()
    try:
        dfs = await loop.run_in_executor(_EXEC, store.load_dataframes)
        if drop_df_keys:
            dfs = {k: v for k, v in dfs.items() if k not in drop_df_keys}
        if not dfs:
            return {"ok": False, "error": "Chat dataset is empty."}

        if kind == "table":
            from code_exec import safe_execute
            from run_chat_local import _build_table_from_result
            exec_out = await loop.run_in_executor(
                _EXEC, lambda: safe_execute(code, dfs, sid=sid))
            if not isinstance(exec_out, dict) or exec_out.get("error"):
                emsg = (exec_out or {}).get("error", "unknown") if isinstance(exec_out, dict) else "unknown"
                log_with_sid(sid, "warning", f"REFRESH_ITEM_TABLE_EXEC_ERROR: {str(emsg)[:200]}",
                             chat_id=chat_id)
                return {"ok": False, "error": "Could not re-run this table with the current data."}
            table = _build_table_from_result(exec_out.get("result"))
            if not table or not (table.get("rows") or []):
                return {"ok": False, "error": "Re-execution did not produce a table."}
            payload = {"ok": True, "kind": "table", "table": table}
            full_table_key = _persist_full_table(store, table, code)
            if full_table_key:
                payload["full_table_key"] = full_table_key
            log_with_sid(sid, "info", "REFRESH_ITEM_OK", kind="table",
                         chat_id=chat_id, rows=table.get("total_rows"))
            return _json_safe(payload)

        # kind == "chart" (default)
        from plot_utils import render_plot_safe
        plot_out = await loop.run_in_executor(
            _EXEC, lambda: render_plot_safe(code, dfs, sid))
        if (not isinstance(plot_out, dict) or plot_out.get("error")
                or plot_out.get("multi_axes")):
            emsg = (plot_out or {}).get("error", "unknown") if isinstance(plot_out, dict) else "unknown"
            log_with_sid(sid, "warning", f"REFRESH_ITEM_CHART_EXEC_ERROR: {str(emsg)[:200]}",
                         chat_id=chat_id)
            return {"ok": False, "error": "Could not re-run this chart with the current data."}
        img = plot_out.get("plotly_html") if plot_out.get("is_plotly") else plot_out.get("image")
        if not img:
            return {"ok": False, "error": "Re-execution did not produce a chart."}
        payload = {"ok": True, "kind": "chart", "image_base64": img,
                   "is_plotly": bool(plot_out.get("is_plotly"))}
        # Fresh chart source data so a subsequent "Show data" shows the
        # refreshed values (same cache + endpoint the live stream uses).
        cd = plot_out.get("chart_data")
        if isinstance(cd, dict) and (cd.get("rows") or cd.get("tables")):
            payload["chart_data_key"] = _cache_full_table(cd)
        log_with_sid(sid, "info", "REFRESH_ITEM_OK", kind="chart",
                     chat_id=chat_id, plotly=bool(plot_out.get("is_plotly")))
        return _json_safe(payload)
    except Exception as e:
        log_with_sid(sid, "error", f"REFRESH_ITEM_ERROR: {type(e).__name__}: {e}",
                     chat_id=chat_id)
        return {"ok": False, "error": "Refresh failed."}


@router.post("/{chat_id}/refresh_item")
async def refresh_item(request: Request, chat_id: str):
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    try:
        body = await request.json()
    except Exception:
        body = {}
    code = ((body or {}).get("code") or "").strip()
    kind = ((body or {}).get("kind") or "chart").strip().lower()
    if not code:
        return JSONResponse({"error": "No code to re-run for this item."}, status_code=400)
    if "###NEXT_PLOT###" in code:
        # Legacy joined multi-chart record — not a single executable block.
        return JSONResponse({"error": "This item cannot be refreshed."}, status_code=400)
    drop, blocked = _role_refresh_block(email, chat_id, code)
    if blocked:
        # Role denial follows the EXECUTION-failure contract (200 {ok:false})
        # — the frontend keeps the previous render and fail-freezes the button.
        log_with_sid(email, "info", "REFRESH_ROLE_DENIED",
                     chat_id=chat_id, tables=blocked)
        return {"ok": False, "code": "ROLE_DENIED", "blocked_tables": blocked,
                "error": "Your role does not include this table's data — "
                         "refresh is unavailable."}
    return await run_item_refresh(chat_id, code, kind, secrets.token_hex(8),
                                  drop_df_keys=drop)


# --- Downloads (chart PNG + table Excel) -------------------------------------
# The B2C dashboard calls these three routes; they were never ported, so every
# "Download" button 404'd. Rendering + Excel build happen locally (Article II:
# raw values stay on the client — nothing here touches the brain).

def _safe_filename(name: str, default: str = "download") -> str:
    """Strip anything that could break a Content-Disposition header or path."""
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", (name or "").strip()).strip("._")
    return cleaned[:120] or default


def _table_to_xlsx_bytes(columns: list, rows: list) -> bytes:
    """Build a .xlsx from the cached/posted table shape ({columns, rows})."""
    import pandas as pd

    cols = list(columns or [])
    data = rows or []
    df = pd.DataFrame(data, columns=cols) if cols else pd.DataFrame(data)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
    return buf.getvalue()


def _df_to_xlsx_bytes(df) -> bytes:
    """Build a .xlsx from a full (uncapped) DataFrame — every row."""
    import pandas as pd

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
    return buf.getvalue()


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@router.post("/{chat_id}/export_plotly_png")
async def export_plotly_png(request: Request, chat_id: str):
    """Render a Plotly chart (raw HTML from the iframe) to a high-res PNG via
    kaleido and stream it back. Body: {html, filename?, scale?}."""
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    try:
        body = await request.json()
    except Exception:
        body = {}
    html = (body or {}).get("html") or ""
    filename = _safe_filename((body or {}).get("filename") or "chart", "chart")
    if not html:
        return JSONResponse({"error": "No chart HTML provided."}, status_code=400)
    try:
        from routes.report import _plotly_html_to_png
        png = _plotly_html_to_png(html, email)
    except Exception as e:
        log_with_sid(email, "error", f"EXPORT_PLOTLY_PNG_FAILED: {e}", chat_id=chat_id)
        png = None
    if not png:
        return JSONResponse({"error": "Could not render chart image."}, status_code=502)
    return Response(
        content=png,
        media_type="image/png",
        headers={"Content-Disposition": f'attachment; filename="{filename}.png"'},
    )


@router.post("/{chat_id}/download_excel/{key}")
async def download_excel(request: Request, chat_id: str, key: str):
    """Build a .xlsx with the FULL (uncapped) result for `key`. Body: {filename?}.

    Re-executes the stored code against the chat's local DataFrames to produce
    the complete result (no 50-row cap), then writes every row. Falls back to
    the stored preview rows only if re-execution errors or yields nothing (e.g.
    chart-data "Show data" records, which have no re-executable code).
    Article II: execution is local; Article IV: guarded, safe fallback."""
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    try:
        body = await request.json()
    except Exception:
        body = {}
    filename = _safe_filename((body or {}).get("filename") or "table", "table")
    store = local_store.ChatDataStore(chat_id)
    rec = _load_full_table_record(store, key)
    if not rec:
        return JSONResponse({"error": "Table not found or expired."}, status_code=404)
    try:
        df = await _reexecute_full_df(chat_id, rec.get("code"), rec.get("result_key"))
        if df is not None and not df.empty:
            xlsx = _df_to_xlsx_bytes(df)
        else:
            xlsx = _table_to_xlsx_bytes(rec.get("columns"), rec.get("rows"))
    except Exception as e:
        log_with_sid(email, "error", f"DOWNLOAD_EXCEL_FAILED: {e}", chat_id=chat_id)
        return JSONResponse({"error": "Could not build Excel file."}, status_code=502)
    return Response(
        content=xlsx,
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


@router.post("/{chat_id}/export_excel")
async def export_excel(request: Request, chat_id: str):
    """Build a .xlsx from the preview table posted verbatim by the frontend.
    Body: {columns, rows, filename?}."""
    email, err = _require_chat(request, chat_id)
    if err:
        return err
    try:
        body = await request.json()
    except Exception:
        body = {}
    filename = _safe_filename((body or {}).get("filename") or "table", "table")
    columns = (body or {}).get("columns") or []
    rows = (body or {}).get("rows") or []
    if not columns and not rows:
        return JSONResponse({"error": "No table data provided."}, status_code=400)
    try:
        xlsx = _table_to_xlsx_bytes(columns, rows)
    except Exception as e:
        log_with_sid(email, "error", f"EXPORT_EXCEL_FAILED: {e}", chat_id=chat_id)
        return JSONResponse({"error": "Could not build Excel file."}, status_code=502)
    return Response(
        content=xlsx,
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )
