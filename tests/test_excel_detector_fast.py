"""Streaming/calamine Excel detection pipeline tests.

Fixtures are built with openpyxl into tmp_path (offline, no brain). They pin
the client-critical contracts across the fast-loader port:
  - multi-row (span=2) headers still flatten correctly,
  - text above the table is still captured for the upload auto-description,
  - hidden/veryHidden sheets are still skipped (commit 6655ebe regression),
  - totals rows are still dropped,
  - a mixed-type column (numbers + error strings) survives as object dtype
    with both value kinds intact (the parquet-fallback contract).
"""
import openpyxl
import pytest

import excel_table_detector as det


def _make_wb(tmp_path, name, sheets):
    """sheets: {sheet_name: {"rows": [...], "state": "visible"|"hidden"|"veryHidden"}}"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sn, spec in sheets.items():
        ws = wb.create_sheet(sn)
        for row in spec["rows"]:
            ws.append(row)
        ws.sheet_state = spec.get("state", "visible")
    path = tmp_path / name
    wb.save(path)
    return path


_SIMPLE_ROWS = [
    ["region", "amount"],
    ["East", 10],
    ["West", 20],
]


def test_single_sheet_simple_table(tmp_path):
    path = _make_wb(tmp_path, "simple.xlsx", {"Data": {"rows": _SIMPLE_ROWS}})
    dfs = det.load_excel_sheets(path, "simple.xlsx")
    assert list(dfs.keys()) == ["simple.xlsx"]  # single visible sheet → bare key
    df = dfs["simple.xlsx"]
    assert list(df.columns) == ["region", "amount"]
    assert len(df) == 2


def test_two_row_header_flattens(tmp_path):
    # S2 gap-fill span signal: the anchor row has a gap where the sub-header
    # row below carries a string → span=2 → flattened "Sales_q1" column.
    rows = [
        ["id", "name", "Sales", None],
        [None, None, "q1", "q2"],
        [1, "amy", 10, 11],
        [2, "bob", 20, 21],
        [3, "cat", 30, 31],
        [4, "dan", 40, 41],
    ]
    path = _make_wb(tmp_path, "span2.xlsx", {"S": {"rows": rows}})
    dfs = det.load_excel_sheets(path, "span2.xlsx")
    df = dfs["span2.xlsx"]
    assert len(df) == 4
    cols = list(df.columns)
    # pandas forward-fills the upper header level, so the gap column under
    # "Sales" flattens to "Sales_q2".
    assert cols == ["id", "name", "Sales_q1", "Sales_q2"]


def test_text_above_table_captured(tmp_path):
    rows = [
        ["Quarterly sales report", None],
        ["Prepared by finance", None],
        ["region", "amount"],
        ["East", 10],
        ["West", 20],
        ["North", 30],
        ["South", 40],
    ]
    det._EXTRACTED_TEXT_ABOVE_TABLE.clear()
    path = _make_wb(tmp_path, "titled.xlsx", {"R": {"rows": rows}})
    dfs = det.load_excel_sheets(path, "titled.xlsx")
    assert len(dfs["titled.xlsx"]) == 4
    captured = det._EXTRACTED_TEXT_ABOVE_TABLE.get("titled.xlsx", "")
    assert "Quarterly sales report" in captured
    assert "Prepared by finance" in captured


def test_hidden_and_veryhidden_sheets_skipped(tmp_path):
    path = _make_wb(tmp_path, "mix.xlsx", {
        "Visible": {"rows": _SIMPLE_ROWS},
        "Hidden": {"rows": _SIMPLE_ROWS, "state": "hidden"},
        "VeryHidden": {"rows": _SIMPLE_ROWS, "state": "veryHidden"},
    })
    dfs = det.load_excel_sheets(path, "mix.xlsx")
    # Only the visible sheet loads AND single-visible keying uses the bare name.
    assert list(dfs.keys()) == ["mix.xlsx"]
    assert len(dfs["mix.xlsx"]) == 2


def test_multi_visible_sheets_keyed_by_sheet(tmp_path):
    path = _make_wb(tmp_path, "multi.xlsx", {
        "A": {"rows": _SIMPLE_ROWS},
        "B": {"rows": [["item", "qty"], ["pen", 5], ["ink", 7]]},
    })
    dfs = det.load_excel_sheets(path, "multi.xlsx")
    assert set(dfs.keys()) == {"multi.xlsx::A", "multi.xlsx::B"}


def test_totals_row_dropped(tmp_path):
    rows = [
        ["region", "amount", "cost"],
        ["East", 10, 5],
        ["West", 20, 8],
        ["Итого", 30, 13],
    ]
    path = _make_wb(tmp_path, "totals.xlsx", {"T": {"rows": rows}})
    dfs = det.load_excel_sheets(path, "totals.xlsx")
    df = dfs["totals.xlsx"]
    assert len(df) == 2
    assert "Итого" not in df["region"].tolist()


def test_mixed_type_column_survives_as_object(tmp_path):
    # Pin the dtype contract the parquet pickle-fallback depends on: numeric
    # values stay numeric, the error strings stay strings, dtype is object.
    rows = [["shop", "disc"]]
    rows += [[f"s{i}", float(i)] for i in range(10)]
    rows += [["sX", "Division by zero"]]
    path = _make_wb(tmp_path, "mixed.xlsx", {"M": {"rows": rows}})
    dfs = det.load_excel_sheets(path, "mixed.xlsx")
    df = dfs["mixed.xlsx"]
    col = df["disc"]
    assert col.dtype == object
    assert "Division by zero" in col.tolist()
    # calamine may load integral floats as ints — either numeric form is fine,
    # what matters is that numbers stay numbers (no string coercion).
    assert any(isinstance(v, (int, float)) and v == 3 for v in col.tolist())


def test_header_near_scan_boundary_keeps_followers(tmp_path):
    # Anchor at row 48 (0-based 47): the density detector's follower rows and
    # the data rows live past row 50 — the HEADER_SCAN_ROWS margin must keep
    # them visible or the sheet silently fails detection.
    rows = [[f"note {i}", None] for i in range(47)]
    rows.append(["region", "amount"])
    rows += [[f"r{i}", i] for i in range(8)]
    path = _make_wb(tmp_path, "deep.xlsx", {"D": {"rows": rows}})
    dfs = det.load_excel_sheets(path, "deep.xlsx")
    assert "deep.xlsx" in dfs
    df = dfs["deep.xlsx"]
    assert list(df.columns) == ["region", "amount"]
    assert len(df) == 8


def test_datetime_headers_become_string_columns(tmp_path):
    # The 12MB customer file's sheet has datetime cells in the HEADER row —
    # raw pd.Timestamp column names crashed the upload's meta write
    # ("keys must be str..., not Timestamp"). Columns must come back as str.
    import datetime as dt
    import json
    rows = [
        ["Brand", dt.datetime(2022, 9, 1), dt.datetime(2022, 10, 1)],
        ["TOUS", 5, 7],
        ["Casio", 3, 4],
        ["Pandora", 1, 2],
    ]
    path = _make_wb(tmp_path, "dt_headers.xlsx", {"B": {"rows": rows}})
    dfs = det.load_excel_sheets(path, "dt_headers.xlsx")
    df = dfs["dt_headers.xlsx"]
    assert all(isinstance(c, str) for c in df.columns)
    # The exact upload shape must be JSON-serializable now.
    json.dumps({c: {"description": "", "values": None} for c in df.columns})


def test_upload_meta_with_hidden_sheets_and_odd_keys(tmp_path, monkeypatch):
    # write_meta must survive non-string schema keys (safety net) AND hidden
    # sheets must never appear in the meta built from detected dataframes.
    import json
    import pandas as pd
    import local_store
    monkeypatch.setattr(local_store.settings, "DATA_ROOT", str(tmp_path))
    local_store._DATAFRAME_CACHE.invalidate()
    path = _make_wb(tmp_path, "mix2.xlsx", {
        "Visible": {"rows": _SIMPLE_ROWS},
        "HiddenData": {"rows": _SIMPLE_ROWS, "state": "hidden"},
    })
    dfs = det.load_excel_sheets(path, "mix2.xlsx")
    store = local_store.UserStore("s_hiddenmeta")
    meta = {"files": [
        {"file_name": k,
         "schema": {"fields": {c: {"description": ""} for c in df.columns}}}
        for k, df in dfs.items()
    ]}
    # Simulate the pre-fix hazard too: a Timestamp key must not crash the write.
    meta["files"][0]["schema"]["fields"][pd.Timestamp("2022-09-01")] = {"description": ""}
    store.write_meta(meta)  # must NOT raise
    saved = json.loads(store.meta_path.read_text(encoding="utf-8"))
    names = [f["file_name"] for f in saved["files"]]
    assert names == ["mix2.xlsx"]  # single visible sheet, no hidden entries
    assert "HiddenData" not in json.dumps(saved)
    assert "2022-09-01" in json.dumps(saved["files"][0]["schema"]["fields"])
    local_store._DATAFRAME_CACHE.invalidate()


def test_listobject_region_still_parsed_heuristically(tmp_path):
    # The ws.tables precheck was dropped with the streaming port; a workbook
    # containing a real ListObject must still parse via the heuristics.
    from openpyxl.worksheet.table import Table, TableStyleInfo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "L"
    for row in _SIMPLE_ROWS:
        ws.append(row)
    tbl = Table(displayName="T1", ref="A1:B3")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tbl)
    path = tmp_path / "listobj.xlsx"
    wb.save(path)
    dfs = det.load_excel_sheets(path, "listobj.xlsx")
    df = dfs["listobj.xlsx"]
    assert list(df.columns) == ["region", "amount"]
    assert len(df) == 2
