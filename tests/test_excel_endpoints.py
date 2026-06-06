"""Unit tests for the client-local Excel download endpoints.

Both endpoints are 100% client-local (no brain call). We assert valid .xlsx
bytes, the filename header, 404 on an expired full-table key, and 401/403
without auth.
"""
import io

import openpyxl
import pytest
from fastapi import FastAPI, Request
from fastapi.testclient import TestClient
from starlette.middleware.sessions import SessionMiddleware

import local_store
import routes.chat as chat_mod

_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
def client(monkeypatch):
    app = FastAPI()
    app.add_middleware(SessionMiddleware, secret_key="test-secret")
    app.include_router(chat_mod.router)

    @app.post("/_login/{email}")
    async def _login(request: Request, email: str):
        request.session["email"] = email
        return {"ok": True}

    return TestClient(app)


def _as_owner(monkeypatch, email="alice@acme.com"):
    monkeypatch.setattr(local_store, "chat_exists", lambda cid: True)
    monkeypatch.setattr(local_store, "get_chat_meta_owner", lambda cid: email)


def test_export_excel_valid_bytes_and_filename(client, monkeypatch):
    _as_owner(monkeypatch)
    client.post("/_login/alice@acme.com")
    resp = client.post(
        "/api/chat/chat123/export_excel",
        json={"columns": ["dept", "salary"],
              "rows": [{"dept": "Eng", "salary": 100}, {"dept": "Sales", "salary": 90}],
              "filename": "table_2026_06_06"},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == _XLSX
    assert 'filename="table_2026_06_06.xlsx"' in resp.headers["content-disposition"]
    # Valid workbook with the posted header + rows.
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["dept", "salary"]
    assert ws.max_row == 3  # header + 2 data rows


def test_export_excel_sanitizes_filename(client, monkeypatch):
    _as_owner(monkeypatch)
    client.post("/_login/alice@acme.com")
    resp = client.post(
        "/api/chat/chat123/export_excel",
        json={"columns": ["a"], "rows": [{"a": 1}], "filename": "../../etc/passwd"},
    )
    assert resp.status_code == 200
    cd = resp.headers["content-disposition"]
    assert "/" not in cd.split("filename=")[1]
    assert ".." not in cd.split("filename=")[1]


def test_download_excel_serves_cached_table(client, monkeypatch):
    _as_owner(monkeypatch)
    client.post("/_login/alice@acme.com")
    key = "deadbeefcafef00d"
    chat_mod._FULL_TABLE_CACHE[key] = {
        "columns": ["x", "y"],
        "rows": [{"x": 1, "y": 2}, {"x": 3, "y": 4}],
        "total_rows": 2,
    }
    try:
        resp = client.post(f"/api/chat/chat123/download_excel/{key}",
                           json={"filename": "full_dump"})
        assert resp.status_code == 200
        assert resp.headers["content-type"] == _XLSX
        assert 'filename="full_dump.xlsx"' in resp.headers["content-disposition"]
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert [c.value for c in wb.active[1]] == ["x", "y"]
    finally:
        chat_mod._FULL_TABLE_CACHE.pop(key, None)


def test_download_excel_expired_key_404(client, monkeypatch):
    _as_owner(monkeypatch)
    client.post("/_login/alice@acme.com")
    resp = client.post("/api/chat/chat123/download_excel/nope_expired",
                       json={"filename": "x"})
    assert resp.status_code == 404
    assert "error" in resp.json()


def test_export_excel_requires_auth_401(client, monkeypatch):
    _as_owner(monkeypatch)
    # No /_login → no session email.
    resp = client.post("/api/chat/chat123/export_excel",
                       json={"columns": ["a"], "rows": [{"a": 1}], "filename": "t"})
    assert resp.status_code == 401


def test_export_excel_non_owner_403(client, monkeypatch):
    # Authenticated as alice, but the chat is owned by bob and not shared.
    monkeypatch.setattr(local_store, "chat_exists", lambda cid: True)
    monkeypatch.setattr(local_store, "get_chat_meta_owner", lambda cid: "bob@acme.com")

    class _Boom:
        def __init__(self, cid):
            raise RuntimeError("no sharing record")

    monkeypatch.setattr(local_store, "ChatDataStore", _Boom)
    client.post("/_login/alice@acme.com")
    resp = client.post("/api/chat/chat123/export_excel",
                       json={"columns": ["a"], "rows": [{"a": 1}], "filename": "t"})
    assert resp.status_code == 403


def test_download_excel_requires_auth_401(client):
    resp = client.post("/api/chat/chat123/download_excel/anykey",
                       json={"filename": "t"})
    assert resp.status_code == 401
