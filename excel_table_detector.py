"""Excel multi-row header + table detection — port of the global streaming
pipeline in the B2C app's `storage.py` (post-calamine rewrite).

Pure Python (openpyxl + calamine + pandas). NO LLM, NO model-based inference.

Pipeline per sheet (no full openpyxl object-graph load anywhere):
  A. Header detection runs on a streamed top slice (`HEADER_SCAN_ROWS` rows,
     openpyxl read_only): BOTH modal-column-width (2A) and string-density
     (2B) anchors, then the 1-vs-2 header-span decision (S2 gap-fill + S3
     sub-header shape; the merged-cell S1 signal and the ws.tables ListObject
     precheck are intentionally NOT used — they need a non-read-only load and
     were validated unnecessary on the global app).
  B. The full data is read ONCE with pd.read_excel(engine="calamine")
     (header=[r] or [r, r+1]); MultiIndex flattened, "Unnamed:" fragments
     dropped, consecutive equal parts deduped.
  C. label_cols auto-detection runs on the LOADED dataframe's head slice.
  D. Totals detection (E = multilingual keyword ∪ F = structural outlier)
     streams over the loaded dataframe with a one-row lookahead; flagged
     total rows are DROPPED from the resulting df AND LOGGED.

Text found above the header is published via `_EXTRACTED_TEXT_ABOVE_TABLE`
so the upload flow can use it as a `file_description` fallback (the header
anchor always sits within the streamed top slice, so the capture is complete).

ENTERPRISE NOTE: This module is the same algorithm as the global B2C app —
only adapted to live next to the client's `local_store.py` (plus the
client-only hidden-sheet skip in `load_excel_sheets`). Any future fix to
header / total detection should be ported back to global to keep the two in
sync. Raw data values still never leave this server.
"""
from __future__ import annotations

import math
import unicodedata
from collections import Counter
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Sequence as _Sequence, Tuple

import pandas as pd


# Stores extracted text found above the table header, keyed by df_name.
# Upload flow reads this to populate file_description in meta.json.
_EXTRACTED_TEXT_ABOVE_TABLE: Dict[str, str] = {}

# Header detection reads only this many top rows (streamed, read-only) instead
# of materializing the whole sheet. The detectors' own scan window is 50 rows
# (sample=50); the +10 margin keeps the density detector's follower check
# (rows[i+1:i+6]), the modal detector's stable_for rows, and the span
# decision's anchor+2 lookahead available for anchors near the bottom of that
# window — the old full-sheet read had those rows, and without the margin a
# header around row ~47-50 would silently lose its span/follower evidence.
# NOTE: global's storage.py uses a bare 50 — port this margin back (module
# policy: keep the two detectors in sync).
HEADER_SCAN_ROWS = 50 + 10


# --- Stage 6 keyword set (multilingual, NFKC-normalized + lowercased) ---
_TOTAL_WORDS: set = {
    # English
    "total", "totals", "grand total", "grand totals", "sum", "subtotal",
    "sub-total", "sub total",
    # Russian
    "итого", "всего", "сумма", "сумм", "итог", "общий итог",
    # German / Italian / Dutch / Finnish / Polish / Czech / Hungarian / Hebrew / etc.
    "summe", "gesamt", "totale", "totaal", "totaali", "yhteensä",
    "totalt", "i alt", "razem", "ogółem", "celkem", "összesen",
    "סה\"כ", "סהכ",
}


def _norm_text(s: str) -> str:
    return unicodedata.normalize("NFKC", s).strip().lower()


def _is_num_cell(v) -> bool:
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return not (isinstance(v, float) and math.isnan(v))
    return False


def _is_str_cell(v) -> bool:
    return isinstance(v, str) and v.strip() != ""


def _nonempty_cells(row: _Sequence) -> list:
    return [v for v in row if v is not None and not (isinstance(v, str) and v.strip() == "")]


def _nonempty_count(row: _Sequence) -> int:
    return sum(1 for v in row if v is not None and not (isinstance(v, str) and v.strip() == ""))


def _native_cell(v):
    """Convert a loaded-dataframe cell back to the value contract openpyxl
    produced, so the existing value-based detectors behave identically:
      - NaN / NaT / pd.NA  -> None   (openpyxl yields None for empty cells; the
        detectors test `v is None`, not pandas-NaN, so this MUST be normalized)
      - numpy scalar       -> native Python int/float/bool (numpy.int64 is NOT
        an `int` subclass, so `_is_num_cell` would otherwise miss integer cells)
    Strings and Python natives pass through unchanged.
    """
    if v is None or isinstance(v, str):
        return v
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    item = getattr(v, "item", None)
    if callable(item):
        try:
            return v.item()
        except Exception:
            return v
    return v


# --- Stage 2A: modal-column-width header anchor ---
def _detect_header_row_modal(rows: List[_Sequence], sample: int = 50,
                              stable_for: int = 3) -> Optional[int]:
    window = rows[:sample]
    counts = [_nonempty_count(r) for r in window]
    meaningful = [c for c in counts if c > 1]
    if not meaningful:
        return None
    modal = Counter(meaningful).most_common(1)[0][0]
    tolerance = max(1, modal - 1)
    for i, c in enumerate(counts):
        if c < modal:
            continue
        followers = counts[i + 1: i + 1 + stable_for]
        if len(followers) < stable_for:
            continue
        if all(fc >= tolerance for fc in followers):
            return i
    return None


# --- Stage 2B: string-density header anchor ---
def _column_type_stable(rows: List[_Sequence], col: int, n_check: int = 5) -> bool:
    seen = []
    for r in rows[:n_check]:
        if col < len(r):
            v = r[col]
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            seen.append("num" if _is_num_cell(v) else "str")
    if not seen:
        return True
    share = max(seen.count("num"), seen.count("str")) / len(seen)
    return share >= 0.6


def _detect_header_row_density(rows: List[_Sequence], sample: int = 50,
                                str_threshold: float = 0.7, min_cells: int = 2,
                                follower_check: int = 5) -> Optional[int]:
    window = rows[:sample]
    for i, r in enumerate(window):
        non_empty = _nonempty_cells(r)
        if len(non_empty) < min_cells:
            continue
        str_frac = sum(1 for v in non_empty if _is_str_cell(v)) / len(non_empty)
        if str_frac < str_threshold:
            continue
        followers = rows[i + 1: i + 1 + follower_check]
        if not followers:
            continue
        n_cols = max(len(r), max((len(f) for f in followers), default=0))
        stable_cols = 0
        data_cols = 0
        for c in range(n_cols):
            col_vals = [f[c] for f in followers
                        if c < len(f) and f[c] is not None
                        and not (isinstance(f[c], str) and f[c].strip() == "")]
            if not col_vals:
                continue
            data_cols += 1
            if _column_type_stable(followers, c):
                stable_cols += 1
        if data_cols == 0:
            continue
        if stable_cols / data_cols >= 0.6:
            return i
    return None


def _pick_header_anchor(rows: List[_Sequence]) -> Tuple[Optional[int], str]:
    idx_a = _detect_header_row_modal(rows, sample=50)
    idx_b = _detect_header_row_density(rows, sample=50)
    candidates: List[Tuple[str, int]] = []
    if idx_b is not None:
        candidates.append(("B", idx_b))
    if idx_a is not None and idx_a != idx_b:
        candidates.append(("A", idx_a))
    if not candidates:
        return None, "none"
    best = None
    best_score = float("inf")
    for label, idx in candidates:
        row = rows[idx]
        non_empty = _nonempty_cells(row)
        if not non_empty:
            continue
        str_frac = sum(1 for v in non_empty if _is_str_cell(v)) / len(non_empty)
        numericish = sum(
            1 for v in non_empty
            if _is_num_cell(v) or (
                isinstance(v, str)
                and v.strip().replace(".", "", 1).replace("-", "", 1).isdigit()
            )
        )
        cell_count_penalty = max(0, 4 - len(non_empty))
        score = (1.0 - str_frac) * 10 + numericish * 5 + cell_count_penalty * 8
        if score < best_score:
            best_score = score
            best = (label, idx)
    return (best[1], best[0]) if best else (None, "none")


# --- Stage 3: header-span auto-decision (1 vs 2) ---
def _detect_header_span(rows: List[_Sequence], anchor_idx: int, ws=None) -> int:
    if anchor_idx + 1 >= len(rows):
        return 1
    upper = rows[anchor_idx]
    lower = rows[anchor_idx + 1]
    upper_nonempty = _nonempty_cells(upper)
    lower_nonempty = _nonempty_cells(lower)
    if not upper_nonempty or not lower_nonempty:
        return 1

    sig1 = False
    if ws is not None:
        anchor_excel_row = anchor_idx + 1
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= anchor_excel_row <= mr.max_row and (mr.max_col - mr.min_col) >= 1:
                sig1 = True
                break

    upper_cells = list(upper)
    lower_cells = list(lower)
    n_cols = max(len(upper_cells), len(lower_cells))
    gap_fill = 0
    for c in range(n_cols):
        u = upper_cells[c] if c < len(upper_cells) else None
        l = lower_cells[c] if c < len(lower_cells) else None
        if (u is None or (isinstance(u, str) and u.strip() == "")) and _is_str_cell(l):
            gap_fill += 1
    sig2 = gap_fill >= 1 and gap_fill >= len(lower_nonempty) * 0.5

    lower_str_frac = sum(1 for v in lower_nonempty if _is_str_cell(v)) / len(lower_nonempty)
    upper_str_frac = sum(1 for v in upper_nonempty if _is_str_cell(v)) / len(upper_nonempty)
    sig3_shape = (
        lower_str_frac >= 0.7
        and upper_str_frac >= 0.7
        and 1 < len(lower_nonempty) < len(upper_nonempty)
    )
    looks_like_data = False
    if anchor_idx + 2 < len(rows):
        d_nonempty = _nonempty_cells(rows[anchor_idx + 2])
        if d_nonempty:
            numeric_present = sum(1 for v in d_nonempty if _is_num_cell(v))
            looks_like_data = numeric_present >= 1 or any(_is_str_cell(v) for v in d_nonempty)
    sig3 = sig3_shape and looks_like_data

    return 2 if (sig1 or sig2 or sig3) else 1


# --- Stage 4: pandas multi-header read + MultiIndex flatten ---
def _read_with_multiheader(
    file_input,
    sheet_name,
    header_row_0based: int,
    header_span: int,
) -> Tuple[List[str], pd.DataFrame]:
    header_rows = list(range(header_row_0based, header_row_0based + header_span))
    if isinstance(file_input, pd.ExcelFile):
        # Reused pre-parsed workbook (engine chosen at ExcelFile construction).
        df = file_input.parse(sheet_name=sheet_name, header=header_rows)
    else:
        df = pd.read_excel(file_input, sheet_name=sheet_name,
                           header=header_rows, engine="calamine")

    def _dedupe(names: List[str]) -> List[str]:
        # Sheets with REPEATED header names (e.g. the same month label twice)
        # otherwise yield duplicate df columns: df[c] then returns a DataFrame
        # (not a Series) and every downstream .dtype/.unique() crashes, and
        # schema dicts keyed by column silently collapse. Pandas' classic
        # mangling convention: second occurrence -> "name.1", third -> ".2".
        seen: dict = {}
        out: List[str] = []
        for n in names:
            if n not in seen:
                seen[n] = 0
                out.append(n)
                continue
            seen[n] += 1
            candidate = f"{n}.{seen[n]}"
            # A generated name can collide with a REAL column ('a','a','a.1')
            # — keep counting until the candidate is genuinely unused.
            while candidate in seen:
                seen[n] += 1
                candidate = f"{n}.{seen[n]}"
            seen[candidate] = 0
            out.append(candidate)
        return out
    flat: List[str] = []
    if isinstance(df.columns, pd.MultiIndex):
        for tup in df.columns:
            parts: List[str] = []
            for part in tup:
                s = "" if part is None else str(part).strip()
                if not s or s.lower().startswith("unnamed"):
                    continue
                parts.append(s)
            deduped: List[str] = []
            for p in parts:
                if not deduped or deduped[-1] != p:
                    deduped.append(p)
            flat.append("_".join(deduped) if deduped else "col")
        flat = _dedupe(flat)
        df.columns = flat
    else:
        flat = [str(c) for c in df.columns]
        # Assign in this branch too: a sheet whose header row holds datetime
        # cells otherwise keeps raw pd.Timestamp COLUMN NAMES, which crash
        # every json.dumps that uses columns as dict keys (meta write on
        # upload → 500). Global's copy computes `flat` here without assigning
        # — same latent bug; port this back.
        flat = _dedupe(flat)
        df.columns = flat
    return flat, df


# --- Stage 5: label_cols auto-detection ---
def _detect_label_cols(rows: List[_Sequence], data_start_idx: int,
                       sample_rows: int = 40, str_threshold: float = 0.6,
                       min_observed: int = 5) -> int:
    if not rows or data_start_idx >= len(rows):
        return 0
    sample = []
    for r in rows[data_start_idx: data_start_idx + sample_rows * 3]:
        if any(v is not None and not (isinstance(v, str) and v.strip() == "") for v in r):
            sample.append(r)
        if len(sample) >= sample_rows:
            break
    if not sample:
        return 0
    n_cols = max(len(r) for r in sample)
    label_cols = 0
    for c in range(n_cols):
        col_vals = [r[c] for r in sample
                    if c < len(r) and r[c] is not None
                    and not (isinstance(r[c], str) and r[c].strip() == "")]
        if len(col_vals) < min_observed:
            break
        str_frac = sum(1 for v in col_vals if _is_str_cell(v)) / len(col_vals)
        if str_frac >= str_threshold:
            label_cols += 1
        else:
            break
    return max(1, label_cols) if label_cols > 0 else 0


# --- Stage 6: total-row detection (E + F) ---
def _first_textual_cell(row: _Sequence) -> Optional[str]:
    for v in row:
        if isinstance(v, str) and v.strip() != "":
            return v
    return None


def _detect_total_by_keyword(row: _Sequence) -> bool:
    text = _first_textual_cell(row)
    if not text:
        return False
    norm = _norm_text(text)
    if norm in _TOTAL_WORDS:
        return True
    for kw in _TOTAL_WORDS:
        if norm.startswith(kw + " ") or norm.startswith(kw + ":") or norm == kw:
            return True
    return False


def _detect_total_by_structure(row: _Sequence, label_cols: int,
                                next_row: Optional[_Sequence] = None,
                                is_last_row: bool = False) -> bool:
    if not row or label_cols <= 0:
        return False
    label_sparse = sum(1 for v in row[:label_cols]
                       if v is None or (isinstance(v, str) and v.strip() == ""))
    label_sparse_frac = label_sparse / max(1, label_cols)
    numeric_after = [v for v in row[label_cols:] if _is_num_cell(v)]
    nonempty_after = [v for v in row[label_cols:]
                      if v is not None and not (isinstance(v, str) and v.strip() == "")]
    if not nonempty_after:
        return False
    numeric_frac = len(numeric_after) / len(nonempty_after)
    strong = label_sparse_frac >= 0.5 and numeric_frac >= 0.7 and len(numeric_after) >= 2
    boost = is_last_row or (next_row is not None and all(
        v is None or (isinstance(v, str) and v.strip() == "") for v in next_row
    ))
    return strong and (boost or label_sparse_frac >= 0.75)


# --- Stage 1: ws.tables precheck (real Excel ListObject) ---
def _read_listobject_table(ws, tbl, raw_rows: List[list]
                            ) -> Optional[Tuple[List[str], pd.DataFrame, int]]:
    try:
        from openpyxl.utils.cell import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
        hdr_count = max(1, int(tbl.headerRowCount or 1))
        totals_count = int(tbl.totalsRowCount or 0)
        header_band_start = min_row - 1
        header_band_end = header_band_start + hdr_count
        data_start = header_band_end
        data_end = max_row - totals_count
        if data_start >= data_end or header_band_end > len(raw_rows):
            return None
        header_band = raw_rows[header_band_start:header_band_end]
        col_names: List[str] = []
        for c in range(min_col - 1, max_col):
            parts: List[str] = []
            for hr in header_band:
                v = hr[c] if c < len(hr) else None
                s = "" if v is None else str(v).strip()
                if not s:
                    continue
                if not parts or parts[-1] != s:
                    parts.append(s)
            col_names.append("_".join(parts) if parts else f"col_{c + 1}")
        data_rows = []
        for r in raw_rows[data_start:data_end]:
            data_rows.append([r[c] if c < len(r) else None for c in range(min_col - 1, max_col)])
        df = pd.DataFrame(data_rows, columns=col_names)
        return col_names, df, hdr_count
    except Exception:
        return None


def _build_text_above(rows: List[list], anchor_idx: int) -> str:
    lines: List[str] = []
    for r in rows[:anchor_idx]:
        non_empty = [c for c in r if c is not None and str(c).strip() != ""]
        if non_empty:
            line_text = " ".join(str(c).strip() for c in non_empty if str(c).strip())
            if line_text:
                lines.append(line_text)
    return "\n".join(lines).strip()


def _read_top_rows(file_input, sheet_name: str, max_rows: int,
                   wb=None) -> Optional[List[list]]:
    """Stream the first `max_rows` rows of VALUES from one sheet using a
    read-only openpyxl load. read_only=True streams rows without building the
    per-cell object graph (the source of the old full-load time/memory blow-up),
    and the early break stops after `max_rows`. Empty cells come back as None,
    matching the value contract the header detectors expect.

    `wb`: an already-open read-only workbook to REUSE (multi-sheet files —
    re-opening per sheet re-parses sharedStrings every time, which dominated
    load time on large workbooks). A caller-provided wb is NOT closed here.
    """
    import openpyxl
    own_wb = None
    try:
        if wb is None:
            wb_input = BytesIO(file_input) if isinstance(file_input, bytes) else str(file_input)
            own_wb = openpyxl.load_workbook(wb_input, read_only=True, data_only=True)
            wb = own_wb
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        rows: List[list] = []
        for r in ws.iter_rows(values_only=True):
            rows.append(list(r))
            if len(rows) >= max_rows:
                break
        return rows
    finally:
        if own_wb is not None:
            try:
                own_wb.close()
            except Exception:
                pass


def _detect_and_extract_table(
    file_input,           # bytes, str path, or Path
    sheet_name: str,
    df_key: str,
    log_prefix: str = "[EXCEL_DETECT]",
    probe_wb=None,        # open read-only openpyxl wb to reuse for the top slice
    excel_file=None,      # open pd.ExcelFile (calamine) to reuse for the body read
) -> Optional[pd.DataFrame]:
    """Detect the table region on one Excel sheet and return the cleaned df.

    Returns the cleaned dataframe (totals dropped) or None on failure. Side
    effects: publishes any captured text above the header into
    `_EXTRACTED_TEXT_ABOVE_TABLE[df_key]`; logs anchor / span / label_cols and
    every dropped total row via `print(..., flush=True)`.

    Pipeline (no longer loads the whole sheet with openpyxl):
      A. Header detection runs on a streamed top-slice (`HEADER_SCAN_ROWS` rows,
         read-only). `_pick_header_anchor` + `_detect_header_span` are reused
         verbatim; `_detect_header_span` is called with ws=None (the merged-cell
         S1 signal simply does not fire — S2/S3 still decide span from values).
      B. The full data is read with the calamine engine via
         `_read_with_multiheader` (header handoff unchanged).
      C. `_detect_label_cols` runs on the LOADED dataframe rows.
      D. `_detect_total_by_keyword` ∪ `_detect_total_by_structure` run over the
         LOADED dataframe rows (same E ∪ F logic, same DROP + LOG behavior).

    NOTE: the `ws.tables` (ListObject) precheck and the merged-cell S1 signal
    are intentionally NOT used on this path — validated unnecessary for this
    data and incompatible with a streaming read.
    """
    # Stage A: streamed header detection over the top slice (values only).
    top_rows = _read_top_rows(file_input, sheet_name, HEADER_SCAN_ROWS, wb=probe_wb)
    if not top_rows:
        return None

    anchor_idx, who = _pick_header_anchor(top_rows)
    if anchor_idx is None:
        print(f"{log_prefix} {df_key}: header anchor NOT FOUND", flush=True)
        return None

    span = _detect_header_span(top_rows, anchor_idx, ws=None)
    print(f"{log_prefix} {df_key}: anchor_row={anchor_idx + 1} (picked-by={who}) span={span}",
          flush=True)

    # Capture text above the header (used by upload flow for file_description)
    text_above = _build_text_above(top_rows, anchor_idx)
    if text_above:
        _EXTRACTED_TEXT_ABOVE_TABLE[df_key] = text_above
        print(f"{log_prefix} {df_key}: captured {len(text_above)} chars of text above table",
              flush=True)

    # Stage B: full data read + flatten (calamine engine, multi-header when
    # span=2). A caller-provided ExcelFile reuses ONE parsed workbook across
    # sheets — pd.read_excel(engine="calamine") on a path re-parses the whole
    # file per call, which dominated load time on multi-sheet workbooks.
    try:
        if excel_file is not None:
            pd_input = excel_file
        else:
            pd_input = BytesIO(file_input) if isinstance(file_input, bytes) else str(file_input)
        flat, df = _read_with_multiheader(pd_input, sheet_name, anchor_idx, span)
    except Exception as e_read:
        print(f"{log_prefix} {df_key}: pandas multi-header read failed "
              f"({type(e_read).__name__}: {e_read})", flush=True)
        return None
    unnamed = sum(1 for c in flat if str(c).lower().startswith("unnamed"))
    print(f"{log_prefix} {df_key}: cols={len(flat)} unnamed={unnamed} data_rows={len(df)}",
          flush=True)

    # Normalize loaded rows to openpyxl's value contract (NaN -> None, numpy
    # scalar -> native) so the existing detectors behave identically. df row k
    # corresponds to sheet row (data_start + k) — pandas discards the preamble
    # and header rows, so the totals row numbering below matches the old path.
    #
    # Rows are normalized ONE AT A TIME (streamed) rather than materialized into
    # a full second list-of-lists: on a large sheet that copy doubled peak
    # memory. The values handed to the detectors — and therefore the resulting
    # drop set — are identical to a full materialization.
    data_start = anchor_idx + span
    n_rows = len(df)

    # Stage C: label_cols auto-detect. `_detect_label_cols` only ever inspects
    # the first `sample_rows * 3` (= 120) rows, so a head slice is equivalent to
    # passing every row.
    head_rows = [[_native_cell(v) for v in r]
                 for r in df.head(40 * 3).itertuples(index=False, name=None)]
    label_cols = _detect_label_cols(head_rows, data_start_idx=0)
    print(f"{log_prefix} {df_key}: label_cols={label_cols}", flush=True)

    # Stage D: totals via E ∪ F over the loaded rows (streamed with a one-row
    # lookahead so `_detect_total_by_structure` still sees `next_row`); DROP + LOG
    flagged: List[Tuple[int, list, str]] = []

    def _scan_total(i: int, row: list, next_row: Optional[list]) -> None:
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
            return
        e = _detect_total_by_keyword(row)
        f = _detect_total_by_structure(
            row,
            label_cols=label_cols,
            next_row=next_row,
            is_last_row=(i == n_rows - 1),
        )
        if e or f:
            flagged.append((i, row, ("E" if e else "") + ("F" if f else "")))

    prev: Optional[Tuple[int, list]] = None  # (index, row) awaiting its next_row
    for i, raw in enumerate(df.itertuples(index=False, name=None)):
        cur = [_native_cell(v) for v in raw]
        if prev is not None:
            _scan_total(prev[0], prev[1], cur)
        prev = (i, cur)
    if prev is not None:
        _scan_total(prev[0], prev[1], None)

    if flagged:
        df_drop_indices: List[int] = []
        for df_i, row_values, tag in flagged:
            if 0 <= df_i < len(df):
                df_drop_indices.append(df_i)
            preview = [str(v)[:30] for v in row_values][:8]
            print(f"{log_prefix} {df_key}: DROPPED total row R{data_start + df_i + 1} [{tag}] {preview}",
                  flush=True)
        if df_drop_indices:
            df = df.drop(df.index[df_drop_indices]).reset_index(drop=True)
        print(f"{log_prefix} {df_key}: dropped {len(df_drop_indices)} total/aggregate row(s)",
              flush=True)

    return df


def _is_valid_table(df: pd.DataFrame, min_rows: int = 1, min_cols: int = 1) -> bool:
    if df is None:
        return False
    if df.shape[0] < min_rows or df.shape[1] < min_cols:
        return False
    unnamed_count = sum(1 for col in df.columns if str(col).startswith("Unnamed"))
    if unnamed_count == len(df.columns):
        return False
    try:
        nan_ratio = df.isna().sum().sum() / (df.shape[0] * df.shape[1])
        if nan_ratio > 0.9:
            return False
    except Exception:
        pass
    return True


def load_excel_sheets(file_path: Path, filename: str) -> Dict[str, pd.DataFrame]:
    """Load every Excel sheet through the validated 6-stage detection pipeline.

    Only VISIBLE sheets are processed — hidden and veryHidden sheets are
    skipped up front, so the single-vs-multi keying below counts visible
    sheets only. Single-visible-sheet workbooks return `{filename: df}`.
    Multi-sheet workbooks return one key per valid sheet:
    `{f"{filename}::{sheet}": df, ...}`. Sheets the pipeline cannot detect a
    table in are skipped. Total/aggregate rows are DROPPED from each df and
    LOGGED.
    """
    import time as _time
    import openpyxl

    total_start = _time.time()
    print(f"[EXCEL] Loading {filename} via validated detection pipeline...", flush=True)

    # ONE read-only openpyxl workbook serves the visibility probe AND every
    # sheet's 50-row top slice; ONE calamine ExcelFile serves every sheet's
    # body read. Re-opening per sheet re-parsed the whole workbook each time
    # (sharedStrings + all sheet XML), which multiplied load time by the
    # sheet count on multi-sheet files.
    wb_probe = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    xl = None
    try:
        try:
            file_size_mb = file_path.stat().st_size / 1024 / 1024
        except Exception:
            file_size_mb = 0.0
        sheet_names: List[str] = []
        for sn in wb_probe.sheetnames:
            try:
                state = wb_probe[sn].sheet_state
            except Exception as e:
                from logger_utils import log_with_sid
                log_with_sid(filename, "warning",
                             f"EXCEL_SHEET_STATE_PROBE_FAILED sheet='{sn}': {e}")
                state = "visible"  # fall back safely: never drop a sheet on probe failure
            if state == "visible":
                sheet_names.append(sn)
            else:
                print(f"[EXCEL]   {filename}::{sn}: SKIPPED (hidden sheet)", flush=True)
        print(f"[EXCEL]   {filename}: {len(sheet_names)} visible sheet(s), {file_size_mb:.2f}MB", flush=True)

        try:
            xl = pd.ExcelFile(str(file_path), engine="calamine")
        except Exception as e:
            from logger_utils import log_with_sid
            log_with_sid(filename, "warning", f"EXCEL_SHARED_HANDLE_FAILED: {e}")
            xl = None  # per-sheet reads fall back to self-opening (slower, same result)

        valid_sheets: List[Tuple[str, pd.DataFrame]] = []
        for sn in sheet_names:
            sheet_start = _time.time()
            df_key = filename if len(sheet_names) == 1 else f"{filename}::{sn}"
            df = _detect_and_extract_table(str(file_path), sn, df_key,
                                           probe_wb=wb_probe, excel_file=xl)
            if _is_valid_table(df):
                valid_sheets.append((sn, df))
                print(f"[EXCEL]   {filename}::{sn}: OK {len(df)} rows in "
                      f"{_time.time() - sheet_start:.3f}s", flush=True)
            else:
                print(f"[EXCEL]   {filename}::{sn}: SKIPPED (pipeline returned no valid table) "
                      f"in {_time.time() - sheet_start:.3f}s", flush=True)
    finally:
        try:
            wb_probe.close()
        except Exception:
            pass
        if xl is not None:
            try:
                xl.close()
            except Exception:
                pass

    dfs: Dict[str, pd.DataFrame] = {}
    if len(valid_sheets) == 1:
        dfs[filename] = valid_sheets[0][1]
    else:
        for sn, df in valid_sheets:
            dfs[f"{filename}::{sn}"] = df

    print(f"[EXCEL] {filename}: TOTAL load_excel_sheets took "
          f"{_time.time() - total_start:.3f}s ({len(dfs)} dataframe(s))", flush=True)
    return dfs
